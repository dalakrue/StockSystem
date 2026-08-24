"""Aggregate Field 3 section for every loaded symbol.

This surface deliberately ignores the active display symbol.

Main features:
    - Middle / Lower aggregate regime views
    - Middle Regime RAS static ALLOW / DISALLOW
    - Strategy Decision: S1 / S2 / S3 / S4 / None
    - Independent Middle Regime historical Finder
    - Finder start date/time + end date/time with free date/time selection
    - Finder range CSV containing one flat table of every record in the selected range
    - Colored XLSX export
    - Persistent Middle Regime history
"""

from __future__ import annotations

from collections.abc import Mapping, MutableMapping
from io import BytesIO
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from core.global_symbol_context import get_global_symbol_context
from ui.field3_mobile_cards_20260722 import (
    is_phone_mode,
    render_responsive_records,
)
from core.regime_age_columns_20260722 import (
    CANDLE_AGE_COLUMN,
    RECENT_CHANGE_RANK_COLUMN,
    REGIME_START_AGE_RANK_COLUMN,
    REGIME_START_STANDARD_COLUMN,
    STANDARD_AGE_COLUMNS,
    enrich_evidence,
    enrich_ranking,
)


# ============================================================================
# DataFrame safety helpers
# ============================================================================

def _sanitize_index_column_collision(frame: Any) -> pd.DataFrame:
    """Return a safe copy when a DataFrame index name duplicates a column.

    Pandas raises ``ValueError: cannot insert <name>, already exists`` when
    ``reset_index()`` is called on a frame whose index is named the same as an
    existing column. Field 3 receives frames from several persisted/cached
    paths, so the index metadata is treated as presentation metadata and is
    cleared whenever it collides with a real column name.
    """
    if not isinstance(frame, pd.DataFrame):
        return pd.DataFrame()
    out = frame.copy()
    # Persisted snapshots can contain repeated labels after concatenation
    # (most commonly ``Datetime``).  Pandas treats a repeated label as a
    # DataFrame instead of a Series and later ``insert``/``reset_index`` calls
    # then fail with the misleading duplicate-column error.  Keep the latest
    # published value for each label before any renderer touches the frame.
    if out.columns.duplicated().any():
        out = out.loc[:, ~out.columns.duplicated(keep="last")].copy()
    index_name = out.index.name
    if index_name is not None and index_name in out.columns:
        out.index = out.index.rename(None)
    return out


def _sanitize_frame_collection(frames: list[pd.DataFrame]) -> list[pd.DataFrame]:
    """Sanitize every persisted frame before concatenation/rendering."""
    return [_sanitize_index_column_collision(frame) for frame in frames
            if isinstance(frame, pd.DataFrame)]


# ============================================================================
# Optional Excel support
# ============================================================================

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Font,
        PatternFill,
    )
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None


# ============================================================================
# Constants
# ============================================================================

AGE_ONLY_RANK_COLUMN = "Regime Age Rank"

# Higher Standard is published separately from the Middle table, but both
# tables use the same canonical ranking snapshot.  Keep the legacy
# ``Regime Age Rank`` alias for older consumers while exposing an explicit
# higher-standard rank/age pair for the new Open/Close sections.
HIGHER_AGE_RANK_COLUMN = "Higher Regime Age Rank"
HIGHER_AGE_DISPLAY_COLUMN = "Higher Candle After Regime Start"

FIELD3_CANONICAL_TABLES_KEY = "field3_canonical_regime_tables_20260823"

MIDDLE_AGE_RANK_COLUMN = "Avg Regime Candle Rank"

# Exact hybrid execution filter requested for the Middle Standard ranking.
# The names intentionally match the supplied pandas reference implementation
# so that CSV/XLSX exports can be consumed by the same backtest/live adapter.
RISK_SCORE_COLUMN = "Risk_Score"
IS_POTENTIAL_BLOCK_COLUMN = "Is_Potential_Block"
HOURLY_BLOCK_COUNT_COLUMN = "Hourly_Block_Count"
FINAL_ACTION_COLUMN = "Final_Action"
RISK_SCORE_THRESHOLD = 1.0
HOURLY_BLOCK_QUOTA = 4
# Maximum weighted-risk SL exits per hour. Remaining exits are ranked by highest risk first.
HOURLY_SL_EXIT_QUOTA = 6

# Weighted regime-age guard.  This replaces the previous historical-loss rank
# with the user-supplied causal score:
#
#     ln(MS**1.5 * CC + 1) / ln(SA + 1)
#
# where MS is Regime Maturity Score, CC is the average completed-regime
# candle-count basis, and SA is the current Middle Standard candle age.
WEIGHTED_RISK_SCORE_COLUMN = "Weighted_Risk_Score"
WEIGHTED_RISK_DECISION_COLUMN = "Weighted Risk Decision"
WEIGHTED_RISK_THRESHOLD = 1.545

# Stop-loss permission derived from the current Middle Standard age (SA).
SL_DECISION_COLUMN = "SL"

MIDDLE_AGE_DISPLAY_COLUMN = (
    "Middle Candle After Regime Start"
)

# Price at the first candle of the current Middle Standard regime.
STARTING_PRICE_COLUMN = "Starting Price"

MIDDLE_AVG_SOURCE_COLUMN = (
    "Middle Avg Regime Candle Count Before Regime Change"
)

AVG_REGIME_CANDLE_COLUMN = (
    "Avg Regime Candle Count Before Regime Change"
)

CANDLE_STATISTICS_COLUMN = (
    "Regime Maturity Score"
)

MIDDLE_COMPLETED_SAMPLE_COLUMN = (
    "Middle Completed Regime Sample Count"
)

COMPLETED_SAMPLE_DISPLAY_COLUMN = (
    "Completed Regime Samples"
)

# --------------------------------------------------------------------------
# RAS
# --------------------------------------------------------------------------

RAS_DECISION_COLUMN = "RAS Decision"

# RAS now intentionally displays only:
#       ALLOW
#       DISALLOW
#
# The score and BUY/SELL explanation are NOT shown in the table.

# --------------------------------------------------------------------------
# Strategy Decision
# --------------------------------------------------------------------------

STRATEGY_DECISION_COLUMN = (
    "Strategy Decision"
)

S1_LABEL = "S1"
S2_LABEL = "S2"
S3_LABEL = "S3"
S4_LABEL = "S4"
NONE_LABEL = "None"

# --------------------------------------------------------------------------
# Historical storage
# --------------------------------------------------------------------------

MIDDLE_HISTORY_STATE_KEY = (
    "field3_middle_regime_history_20260722"
)

MIDDLE_HISTORY_FILE = Path(
    "data/field3_middle_regime_history_20260722.parquet"
)
MIDDLE_HISTORY_CSV_FILE = Path(
    "data/field3_middle_regime_history_20260722.csv"
)

# --------------------------------------------------------------------------
# Finder
# --------------------------------------------------------------------------

FINDER_RESULT_STATE_KEY = (
    "field3_middle_finder_result_20260816"
)

FINDER_START_DAY_STATE_KEY = (
    "field3_middle_finder_start_day_20260816"
)

FINDER_START_TIME_STATE_KEY = (
    "field3_middle_finder_start_time_20260816"
)

FINDER_END_DAY_STATE_KEY = (
    "field3_middle_finder_end_day_20260816"
)

FINDER_END_TIME_STATE_KEY = (
    "field3_middle_finder_end_time_20260816"
)

FINDER_LAST_RANGE_STATE_KEY = (
    "field3_middle_finder_last_range_20260816"
)

FINDER_VIEW_MODE_STATE_KEY = (
    "field3_middle_finder_view_mode_20260818"
)

FINDER_PAGE_SIZE_STATE_KEY = (
    "field3_middle_finder_page_size_20260818"
)


# --------------------------------------------------------------------------
# Field Open / Close state
# --------------------------------------------------------------------------

FIELD_CLOSE_HIGHER_RANKING_KEY = (
    "field_close_higher_standard_regime_ranking"
)

FIELD_OPEN_HIGHER_RANKING_KEY = (
    "field_open_higher_standard_regime_ranking"
)

FIELD_CLOSE_HIGHER_MOBILE_KEY = (
    "field_close_higher_standard_regime_ranking_mobile"
)

FIELD_OPEN_HIGHER_MOBILE_KEY = (
    "field_open_higher_standard_regime_ranking_mobile"
)

FIELD_SECTION_ORDER_KEY = (
    "field3_field_section_order"
)

MIDDLE_NUMERIC_HIGHLIGHT_COLUMNS = [
    CANDLE_AGE_COLUMN,
    STARTING_PRICE_COLUMN,
    CANDLE_STATISTICS_COLUMN,
    RISK_SCORE_COLUMN,
    HOURLY_BLOCK_COUNT_COLUMN,
    WEIGHTED_RISK_SCORE_COLUMN,
    AVG_REGIME_CANDLE_COLUMN,
    COMPLETED_SAMPLE_DISPLAY_COLUMN,
]

STANDARD_LABELS = {
    "LOWER": "Lower Standard Summary",
    "MIDDLE": "Middle Standard Summary",
}


# ============================================================================
# General helpers
# ============================================================================

def _load_saved_if_needed(
    state: MutableMapping[str, Any],
) -> None:

    ranking = state.get(
        "field3_multisymbol_regime_20260708"
    )

    evidence = state.get(
        "field3_regime_evidence_v2"
    )

    if (
        isinstance(ranking, pd.DataFrame)
        and not ranking.empty
        and isinstance(evidence, pd.DataFrame)
        and not evidence.empty
    ):
        return

    try:

        from core.field3_three_regime_engine import (
            load_saved_field3_v2,
        )

        load_saved_field3_v2(state)

    except Exception as exc:

        state[
            "field3_summary_reload_error_20260722"
        ] = (
            f"{type(exc).__name__}: {exc}"
        )


def _scope_to_loaded(
    frame: pd.DataFrame,
    loaded: list[str],
) -> pd.DataFrame:

    if (
        not isinstance(
            frame,
            pd.DataFrame,
        )
        or frame.empty
        or "Symbol" not in frame.columns
        or not loaded
    ):
        return frame.copy()

    allowed = {
        str(symbol).upper()
        for symbol in loaded
    }

    return frame.loc[
        frame["Symbol"]
        .astype(str)
        .str.upper()
        .isin(allowed)
    ].copy()


def _fill_middle_starting_prices_from_loaded_candles(
    ranking: pd.DataFrame,
    state: MutableMapping[str, Any],
) -> pd.DataFrame:
    """Backfill Starting Price for older saved rankings when candles exist.

    New calculation generations persist the value directly. This fallback
    keeps an older saved ranking useful by reading the already-loaded exact
    symbol candles, without fetching data or changing any calculation.
    """
    if not isinstance(ranking, pd.DataFrame) or ranking.empty or "Symbol" not in ranking.columns:
        return ranking.copy() if isinstance(ranking, pd.DataFrame) else pd.DataFrame()
    out = ranking.copy()
    if STARTING_PRICE_COLUMN not in out.columns:
        out[STARTING_PRICE_COLUMN] = pd.Series(pd.NA, index=out.index, dtype="Float64")
    prices = pd.to_numeric(out[STARTING_PRICE_COLUMN], errors="coerce")
    if prices.notna().all():
        out[STARTING_PRICE_COLUMN] = prices.round(5)
        return out
    frames = state.get("canonical_symbol_candles") if isinstance(state, Mapping) else None
    if not isinstance(frames, Mapping):
        out[STARTING_PRICE_COLUMN] = prices.round(5)
        return out
    age_col = _first_existing_column(
        out,
        ("Middle Candle After Regime Start", "Middle Regime Age", "Regime Age"),
    )
    if age_col is None:
        out[STARTING_PRICE_COLUMN] = prices.round(5)
        return out
    for idx, row in out.iterrows():
        if pd.notna(prices.loc[idx]):
            continue
        symbol = str(row.get("Symbol") or "").strip().upper()
        frame = frames.get(symbol)
        if not isinstance(frame, pd.DataFrame):
            frame = frames.get(symbol.replace("/", ""))
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            continue
        age = pd.to_numeric(pd.Series([row.get(age_col)]), errors="coerce").iloc[0]
        if pd.isna(age) or float(age) < 1:
            continue
        source = next((c for c in ("open", "Open", "close", "Close") if c in frame.columns), None)
        if source is None:
            continue
        try:
            position = max(0, len(frame) - int(round(float(age))))
            value = float(pd.to_numeric(frame[source], errors="coerce").iloc[position])
            if np.isfinite(value):
                prices.loc[idx] = value
        except Exception:
            continue
    out[STARTING_PRICE_COLUMN] = prices.round(5)
    return out


def _first_existing_column(
    frame: pd.DataFrame,
    candidates: tuple[str, ...],
) -> str | None:

    if not isinstance(
        frame,
        pd.DataFrame,
    ):
        return None

    for column in candidates:
        if column in frame.columns:
            return column

    normalized = {
        str(column)
        .strip()
        .upper(): column
        for column in frame.columns
    }

    for candidate in candidates:

        found = normalized.get(
            str(candidate)
            .strip()
            .upper()
        )

        if found is not None:
            return found

    return None


# ============================================================================
# Bias normalization
# ============================================================================

def _normalize_bias(
    value: Any,
) -> str:

    if value is None:
        return "NEUTRAL"

    try:
        if pd.isna(value):
            return "NEUTRAL"
    except Exception:
        pass

    normalized = (
        str(value)
        .strip()
        .upper()
    )

    if normalized in {
        "BUY",
        "SELL",
        "NEUTRAL",
    }:
        return normalized

    if normalized in {
        "BULLISH",
        "LONG",
        "UP",
        "STRONG BUY",
    }:
        return "BUY"

    if normalized in {
        "BEARISH",
        "SHORT",
        "DOWN",
        "STRONG SELL",
    }:
        return "SELL"

    return "NEUTRAL"


# ============================================================================
# RAS calculation
# ============================================================================

def _ras_score_for_direction(
    hrb: str,
    mrb: str,
    lrb: str,
    direction: str,
) -> int:

    score = 0

    # Higher regime weight = 2
    if hrb == direction:
        score += 2
    elif hrb in {
        "BUY",
        "SELL",
    }:
        score -= 2

    # Middle regime weight = 1
    if mrb == direction:
        score += 1
    elif mrb in {
        "BUY",
        "SELL",
    }:
        score -= 1

    # Lower regime weight = 1
    if lrb == direction:
        score += 1
    elif lrb in {
        "BUY",
        "SELL",
    }:
        score -= 1

    return score


def get_ras_status(
    row: pd.Series,
) -> str:
    """
    Return ONLY:

        ALLOW
        DISALLOW

    A direction is valid when:
        score >= 1
        AND Higher Regime Bias does not oppose it.
    """

    lrb = _normalize_bias(
        row.get(
            "Lower Regime Bias"
        )
    )

    mrb = _normalize_bias(
        row.get(
            "Middle Regime Bias"
        )
    )

    hrb = _normalize_bias(
        row.get(
            "Higher Standard Regime Bias"
        )
    )

    buy_score = _ras_score_for_direction(
        hrb=hrb,
        mrb=mrb,
        lrb=lrb,
        direction="BUY",
    )

    sell_score = _ras_score_for_direction(
        hrb=hrb,
        mrb=mrb,
        lrb=lrb,
        direction="SELL",
    )

    buy_valid = (
        buy_score >= 1
        and hrb != "SELL"
    )

    sell_valid = (
        sell_score >= 1
        and hrb != "BUY"
    )

    if buy_valid or sell_valid:
        return "ALLOW"

    return "DISALLOW"


def _add_ras_decision_column(
    frame: pd.DataFrame,
) -> pd.DataFrame:

    if not isinstance(
        frame,
        pd.DataFrame,
    ):
        return pd.DataFrame()

    result = frame.copy()

    if result.empty:

        result[
            RAS_DECISION_COLUMN
        ] = pd.Series(
            dtype="object"
        )

        return result

    required_columns = {
        "Lower Regime Bias",
        "Middle Regime Bias",
        "Higher Standard Regime Bias",
    }

    if not required_columns.issubset(
        set(result.columns)
    ):

        result[
            RAS_DECISION_COLUMN
        ] = "DISALLOW"

        return result

    result[
        RAS_DECISION_COLUMN
    ] = result.apply(
        get_ras_status,
        axis=1,
    )

    return result


# ============================================================================
# Strategy Decision
# ============================================================================

def _resolve_strategy_source_columns(
    frame: pd.DataFrame,
) -> dict[str, str | None]:

    return {
        "START_AGE": _first_existing_column(
            frame,
            (
                CANDLE_AGE_COLUMN,
                "Starting Age",
                "Regime Age",
                "Candle After Regime Start",
            ),
        ),
        "LRB": _first_existing_column(
            frame,
            (
                "Lower Regime Bias",
                "LRB",
                "Lower Bias",
            ),
        ),
        "MRB": _first_existing_column(
            frame,
            (
                "Middle Regime Bias",
                "MRB",
                "Middle Bias",
            ),
        ),
        "HRB": _first_existing_column(
            frame,
            (
                "Higher Standard Regime Bias",
                "Higher Regime Bias",
                "HRB",
                "Higher Bias",
            ),
        ),
    }


def _numeric_row_value(
    row: pd.Series,
    column: str | None,
) -> float | None:

    if column is None:
        return None

    value = pd.to_numeric(
        row.get(column),
        errors="coerce",
    )

    if pd.isna(value):
        return None

    return float(value)


def _row_bias(
    row: pd.Series,
    column: str | None,
) -> str:

    if column is None:
        return "NEUTRAL"

    return _normalize_bias(
        row.get(column)
    )


def get_strategy_decision(
    row: pd.Series,
) -> str:
    """Return the corrected strategy decision and its direction text."""

    ras = str(
        row.get(
            RAS_DECISION_COLUMN,
            "DISALLOW",
        )
    ).strip().upper()

    columns = _resolve_strategy_source_columns(
        row.to_frame().T
    )

    start_age = _numeric_row_value(
        row,
        columns["START_AGE"],
    )
    lrb = _row_bias(row, columns["LRB"])
    mrb = _row_bias(row, columns["MRB"])
    hrb = _row_bias(row, columns["HRB"])
    directional = {"BUY", "SELL"}

    def _label(direction: str, strategy: str) -> str:
        # A strategy without a directional BUY/SELL bias is a true None state.
        return f"{direction} — {strategy}" if direction in directional else NONE_LABEL

    # S4 is only valid after SA is above 15. Lower ages continue through
    # normal strategy evaluation instead of creating an S4 exit state.
    if ras != "ALLOW" and start_age is not None and start_age > 15:
        return _label(hrb, S4_LABEL)

    if start_age is None:
        return NONE_LABEL

    # S1: SA below 15 and (HRB == MRB or MRB is neutral).
    if (
        start_age < 15
        and (
            hrb == mrb
            or mrb == "NEUTRAL"
        )
    ):
        return _label(hrb, S1_LABEL)

    # S2: Starting Age > 20, only LRB + MRB match, HRB differs.
    # Valid examples: BUY/BUY/SELL, BUY/BUY/NEUTRAL,
    #                 SELL/SELL/BUY, SELL/SELL/NEUTRAL.
    if (
        start_age > 20
        and lrb == mrb
        and lrb in directional
        and hrb != lrb
    ):
        return _label(mrb, S2_LABEL)

    # S3: Starting Age > 20 and MRB + HRB match. LRB is irrelevant.
    if (
        start_age > 20
        and mrb == hrb
        and mrb in directional
    ):
        opposite_mrb = "SELL" if mrb == "BUY" else "BUY"
        return _label(opposite_mrb, S3_LABEL)

    return NONE_LABEL


def _add_strategy_decision_column(
    frame: pd.DataFrame,
) -> pd.DataFrame:

    if not isinstance(
        frame,
        pd.DataFrame,
    ):
        return pd.DataFrame()

    result = frame.copy()

    if result.empty:

        result[
            STRATEGY_DECISION_COLUMN
        ] = pd.Series(
            dtype="object"
        )

        return result

    # Always calculate RAS first.
    result = _add_ras_decision_column(
        result
    )

    result[
        STRATEGY_DECISION_COLUMN
    ] = result.apply(
        get_strategy_decision,
        axis=1,
    )

    return result


def _first_numeric_column(frame: pd.DataFrame, names: tuple[str, ...]) -> str | None:
    """Return the first present column that contains at least one number."""
    if not isinstance(frame, pd.DataFrame):
        return None
    for name in names:
        if name not in frame.columns:
            continue
        values = pd.to_numeric(frame[name], errors="coerce")
        if values.notna().any():
            return name
    return None


def _add_weighted_risk_score(
    table: pd.DataFrame,
    state: Mapping[str, Any] | None = None,
) -> pd.DataFrame:
    """Add the supplied MS/CC/SA weighted risk score and decision.

    Formula::

        Weighted_Risk_Score = ln(MS**1.5 * CC + 1) / ln(SA + 1)

    A score at or above ``1.545`` is a ``BLOCK``.  Missing/invalid inputs are
    deliberately treated as ``ALLOW`` so an incomplete snapshot is never
    silently converted into a hard trading veto.  The score is still exposed
    as ``NaN`` in that case for transparent diagnostics.
    """
    if not isinstance(table, pd.DataFrame):
        return pd.DataFrame()

    result = table.copy()
    if result.empty:
        result[WEIGHTED_RISK_SCORE_COLUMN] = pd.Series(dtype="Float64")
        result[WEIGHTED_RISK_DECISION_COLUMN] = pd.Series(dtype="object")
        result[SL_DECISION_COLUMN] = pd.Series(dtype="object")
        return _add_hourly_capped_risk_filter(result, state)

    # MS = Regime Maturity Score; CC = completed-regime candle-count basis;
    # SA = current Middle Standard candle age.  Accept legacy aliases so old
    # saved generations gain the new column without a recalculation.
    ms_column = _first_numeric_column(
        result,
        (
            CANDLE_STATISTICS_COLUMN,
            "MS (Regime Maturity Score)",
            "MS",
            "Middle Regime Maturity Score",
        ),
    )
    cc_column = _first_numeric_column(
        result,
        (
            AVG_REGIME_CANDLE_COLUMN,
            MIDDLE_AVG_SOURCE_COLUMN,
            "Candle Count Before Regime Change",
            "CC (Candle Count Before Regime Change)",
            "CC",
        ),
    )
    sa_column = _first_numeric_column(
        result,
        (
            CANDLE_AGE_COLUMN,
            "Starting Age",
            "SA (Starting Age)",
            "SA",
        ),
    )

    ms = pd.to_numeric(result[ms_column], errors="coerce") if ms_column else pd.Series(np.nan, index=result.index)
    cc = pd.to_numeric(result[cc_column], errors="coerce") if cc_column else pd.Series(np.nan, index=result.index)
    sa = pd.to_numeric(result[sa_column], errors="coerce") if sa_column else pd.Series(np.nan, index=result.index)
    valid = ms.ge(0) & cc.ge(0) & sa.gt(0)
    numerator = np.log(np.maximum(ms.clip(lower=0).to_numpy(dtype=float) ** 1.5 * cc.clip(lower=0).to_numpy(dtype=float) + 1.0, 1.0))
    denominator = np.log(sa.clip(lower=0).to_numpy(dtype=float) + 1.0)
    valid_array = valid.fillna(False).to_numpy(dtype=bool)
    score_values = np.full(len(result), np.nan, dtype=float)
    safe_denominator = valid_array & np.isfinite(denominator) & (denominator > 0)
    np.divide(numerator, denominator, out=score_values, where=safe_denominator)
    score = pd.Series(score_values, index=result.index, dtype="Float64")
    result[WEIGHTED_RISK_SCORE_COLUMN] = score.round(4)
    result[WEIGHTED_RISK_DECISION_COLUMN] = np.where(score.ge(WEIGHTED_RISK_THRESHOLD).fillna(False), "BLOCK", "ALLOW")

    # Keep the existing SL display.  The legacy weighted score remains a
    # diagnostic column, while the exact capped MS/CC/SA rule below is the
    # authoritative execution gate for this table.
    result[SL_DECISION_COLUMN] = np.where(sa < 15, "No SL", "Allowed")
    result = _add_ras_decision_column(result)
    result = _add_hourly_capped_risk_filter(result, state)
    # Apply the four-per-hour priority cap to the public decision column too.
    # Final_Action is retained as a diagnostic/execution alias, but rows that
    # lose the quota must not continue to display BLOCK in Weighted Risk Decision.
    if FINAL_ACTION_COLUMN in result.columns:
        result[WEIGHTED_RISK_DECISION_COLUMN] = np.where(
            result[FINAL_ACTION_COLUMN].astype(str).eq("BLOCK"), "BLOCK", "ALLOW"
        )
    result.loc[result[FINAL_ACTION_COLUMN].eq("BLOCK"), RAS_DECISION_COLUMN] = "DISALLOW"

    # Exit protection: block decisions and a direction flip from the previous
    # hourly snapshot of the same symbol invalidate the current trade bias.
    if isinstance(state, Mapping):
        history = state.get(MIDDLE_HISTORY_STATE_KEY)
        if isinstance(history, pd.DataFrame) and not history.empty and "Symbol" in result.columns:
            hist = history.copy()
            if "Datetime" in hist.columns:
                hist["Datetime"] = pd.to_datetime(hist["Datetime"], errors="coerce")
            if STRATEGY_DECISION_COLUMN in hist.columns:
                for idx, row in result.iterrows():
                    current_symbol = row.get("Symbol")
                    previous = hist[hist["Symbol"].eq(current_symbol)]
                    if "Datetime" in previous.columns and "Datetime" in result.columns:
                        previous = previous[previous["Datetime"] < pd.to_datetime(row.get("Datetime"), errors="coerce")]
                    if not previous.empty:
                        old_bias = str(previous.iloc[-1].get(STRATEGY_DECISION_COLUMN, "")).split("—")[0].strip()
                        new_bias = str(row.get(STRATEGY_DECISION_COLUMN, "")).split("—")[0].strip()
                        if old_bias in {"BUY", "SELL"} and new_bias in {"BUY", "SELL"} and old_bias != new_bias:
                            result.at[idx, SL_DECISION_COLUMN] = "Exit (Bias Change)"
    result.loc[result[WEIGHTED_RISK_DECISION_COLUMN].eq("BLOCK"), SL_DECISION_COLUMN] = "Exit (Risk Block)"
    result[STRATEGY_DECISION_COLUMN] = result.apply(get_strategy_decision, axis=1)
    result = _apply_hourly_sl_exit_cap(result)
    retired_columns = [
        column for column in result.columns
        if str(column).strip().lower() in {"historical_loss_pattern_risk_rank", "historicalriskrank"}
    ]
    return result.drop(
        columns=retired_columns + ["_HistoricalRiskScore", "_HistoricalRiskHour", "HistoricalRiskHour"],
        errors="ignore",
    )



def _apply_hourly_sl_exit_cap(table: pd.DataFrame) -> pd.DataFrame:
    """Limit SL exits to six per hour and prioritize highest Weighted_Risk_Score."""
    if not isinstance(table, pd.DataFrame) or table.empty or SL_DECISION_COLUMN not in table.columns:
        return table
    result = table.copy()
    dt_col = _find_datetime_column(result)
    if dt_col:
        hours = _normalize_datetime_series(result[dt_col]).dt.floor("h")
    else:
        hours = pd.Series("CURRENT_HOUR", index=result.index)
    result["_SL_Hour"] = hours.astype("object").where(hours.notna(), "CURRENT_HOUR")
    exit_mask = result[SL_DECISION_COLUMN].astype(str).str.contains("Exit", case=False, na=False)
    if WEIGHTED_RISK_SCORE_COLUMN in result.columns:
        priority = pd.to_numeric(result[WEIGHTED_RISK_SCORE_COLUMN], errors="coerce").fillna(-1)
    else:
        priority = pd.Series(-1, index=result.index)
    result["_SL_Priority"] = priority
    keep = pd.Series(False, index=result.index)
    for _, group in result[exit_mask].groupby("_SL_Hour", dropna=False):
        chosen = group.sort_values("_SL_Priority", ascending=False, kind="mergesort").head(HOURLY_SL_EXIT_QUOTA)
        keep.loc[chosen.index] = True
    result.loc[exit_mask & ~keep, SL_DECISION_COLUMN] = "Hold (SL Exit Hourly Cap)"
    return result.drop(columns=["_SL_Hour", "_SL_Priority"], errors="ignore")


def _add_hourly_capped_risk_filter(
    table: pd.DataFrame,
    state: Mapping[str, Any] | None = None,
) -> pd.DataFrame:
    """Apply the exact ``MS * CC / SA`` rule with a four-block hourly cap.

    A potential block is counted only when it is actually blocked.  Once four
    blocks have been issued in a calendar hour, later potential blocks in that
    hour are allowed, exactly as the requested execution rule specifies.  The
    calculation is deterministic for a historical table: rows are evaluated
    chronologically (with the incoming row order as a stable tie-breaker) and
    then returned in their original order.
    """
    del state  # reserved for a live adapter; historical tables are self-contained
    if not isinstance(table, pd.DataFrame):
        return pd.DataFrame()

    result = table.copy()
    if result.empty:
        result[RISK_SCORE_COLUMN] = pd.Series(dtype="Float64")
        result[IS_POTENTIAL_BLOCK_COLUMN] = pd.Series(dtype="boolean")
        result[HOURLY_BLOCK_COUNT_COLUMN] = pd.Series(dtype="Int64")
        result[FINAL_ACTION_COLUMN] = pd.Series(dtype="object")
        return result

    ms_column = _first_numeric_column(
        result,
        (
            CANDLE_STATISTICS_COLUMN,
            "MS (Regime Maturity Score)",
            "MS",
            "Middle Regime Maturity Score",
        ),
    )
    cc_column = _first_numeric_column(
        result,
        (
            AVG_REGIME_CANDLE_COLUMN,
            MIDDLE_AVG_SOURCE_COLUMN,
            "Candle Count Before Regime Change",
            "CC (Candle Count Before Regime Change)",
            "CC",
        ),
    )
    sa_column = _first_numeric_column(
        result,
        (
            CANDLE_AGE_COLUMN,
            "Starting Age",
            "SA (Starting Age)",
            "SA",
        ),
    )

    ms = (
        pd.to_numeric(result[ms_column], errors="coerce")
        if ms_column
        else pd.Series(np.nan, index=result.index)
    )
    cc = (
        pd.to_numeric(result[cc_column], errors="coerce")
        if cc_column
        else pd.Series(np.nan, index=result.index)
    )
    sa = (
        pd.to_numeric(result[sa_column], errors="coerce")
        if sa_column
        else pd.Series(np.nan, index=result.index)
    )

    valid = ms.ge(0) & cc.ge(0) & sa.gt(0)
    with np.errstate(divide="ignore", invalid="ignore"):
        scores = (ms * cc).div(sa)
    scores = scores.where(valid & np.isfinite(scores), np.nan)
    result[RISK_SCORE_COLUMN] = scores.round(6).astype("Float64")
    potential = result[RISK_SCORE_COLUMN].ge(RISK_SCORE_THRESHOLD).fillna(False).astype(bool)
    result[IS_POTENTIAL_BLOCK_COLUMN] = potential.astype("boolean")

    # ``Completed Candle`` is the standard table's timestamp.  Support the
    # broader aliases used by historical exports and live adapters as well.
    datetime_column = _find_datetime_column(result)
    if datetime_column is not None:
        timestamps = _normalize_datetime_series(result[datetime_column])
    else:
        timestamps = pd.Series(pd.NaT, index=result.index, dtype="datetime64[ns]")

    # A current ranking can legitimately omit a timestamp.  Treat all such
    # rows as one current-hour bucket instead of silently resetting the quota
    # for every symbol.
    hour_values = timestamps.dt.floor("h")
    hour_values = hour_values.astype("object").where(hour_values.notna(), "CURRENT_HOUR")
    result["Hour"] = hour_values

    working = pd.DataFrame(
        {
            "__original_order": np.arange(len(result), dtype=int),
            "__timestamp": timestamps,
            "__hour": hour_values,
            "__potential": potential.to_numpy(dtype=bool),
        },
        index=result.index,
    )
    working = working.sort_values(
        ["__timestamp", "__original_order"],
        na_position="last",
        kind="mergesort",
    )

    # Priority rule: when more than four symbols qualify in the same hour,
    # keep only the four largest Risk_Score values. This prevents row order
    # from deciding which symbols receive protection.
    working["__risk_score"] = result[RISK_SCORE_COLUMN].reindex(working.index)
    working["__selected"] = False
    for hour_key, group in working.groupby("__hour", dropna=False):
        candidates = group[group["__potential"]].sort_values(
            ["__risk_score", "__original_order"],
            ascending=[False, True],
            kind="mergesort",
        )
        working.loc[candidates.head(HOURLY_BLOCK_QUOTA).index, "__selected"] = True

    counts: dict[Any, int] = {}
    actions: dict[Any, str] = {}
    for row_index, row in working.iterrows():
        hour_key = row["__hour"]
        if bool(row["__selected"]):
            actions[row_index] = "BLOCK"
        else:
            actions[row_index] = "ALLOW"
        counts[row_index] = int(working.loc[(working["__hour"] == hour_key) & working["__selected"]].shape[0])

    result[HOURLY_BLOCK_COUNT_COLUMN] = pd.Series(
        [counts.get(index, 0) for index in result.index],
        index=result.index,
        dtype="Int64",
    )
    result[FINAL_ACTION_COLUMN] = pd.Series(
        [actions.get(index, "ALLOW") for index in result.index],
        index=result.index,
        dtype="object",
    )
    return result
# ============================================================================
# Standard table
# ============================================================================

def _standard_table(
    evidence: pd.DataFrame,
    ranking: pd.DataFrame,
    standard: str,
) -> pd.DataFrame:

    standard_series = evidence.get(
        "Standard",
        pd.Series(
            "",
            index=evidence.index,
        ),
    )

    scoped = evidence.loc[
        standard_series
        .astype(str)
        .str.upper()
        .eq(
            standard
        )
    ].copy()

    if scoped.empty:
        return scoped

    current_rank = ranking[
        [
            c
            for c in (
                "Symbol",
                "Rank",
            )
            if c in ranking.columns
        ]
    ].copy()

    if set(
        current_rank.columns
    ) == {
        "Symbol",
        "Rank",
    }:

        scoped = scoped.merge(
            current_rank.rename(
                columns={
                    "Rank": "Final Rank"
                }
            ),
            on="Symbol",
            how="left",
        )

    if CANDLE_AGE_COLUMN in scoped.columns:

        age = pd.to_numeric(
            scoped[
                CANDLE_AGE_COLUMN
            ],
            errors="coerce",
        )

        scoped[
            REGIME_START_AGE_RANK_COLUMN
        ] = (
            age.rank(
                method="min",
                ascending=False,
                na_option="bottom",
            )
            .astype("Int64")
        )

        scoped[
            RECENT_CHANGE_RANK_COLUMN
        ] = (
            age.rank(
                method="min",
                ascending=True,
                na_option="bottom",
            )
            .astype("Int64")
        )

    columns = [
        REGIME_START_AGE_RANK_COLUMN,
        RECENT_CHANGE_RANK_COLUMN,
        "Final Rank",
        "Symbol",
        "Regime State",
        "Bias",
        "Posterior Probability",
        "Persistence Probability",
        "Expected Duration",
        CANDLE_AGE_COLUMN,
        "Changepoint Probability",
        "Transition Risk",
        "Calibrated Reliability",
        "Sample Count",
        "Data Quality Grade",
        "Completed Candle",
    ]

    table = scoped[
        [
            c
            for c in columns
            if c in scoped.columns
        ]
    ].copy()

    sort_cols = [
        c
        for c in (
            RECENT_CHANGE_RANK_COLUMN,
            "Final Rank",
            "Symbol",
        )
        if c in table.columns
    ]

    if sort_cols:

        return (
            table
            .sort_values(
                sort_cols,
                kind="mergesort",
            )
            .reset_index(
                drop=True
            )
        )

    return table.reset_index(
        drop=True
    )


# ============================================================================
# Higher regime ranking
# ============================================================================

def _age_only_ranking(
    ranking: pd.DataFrame,
) -> pd.DataFrame:
    """Build the Higher Standard ranking from the same snapshot as Middle.

    This used to be a dead compatibility helper: it calculated a frame but
    the main renderer never called it, so opening/closing the higher section
    could not show anything.  The function now keeps a dedicated higher age
    column/rank and adds the same RAS/Strategy columns used by Middle.  No
    calculation or provider call is performed here; ``ranking`` is already
    the enriched, saved Field 3 snapshot.
    """

    if not isinstance(ranking, pd.DataFrame) or ranking.empty:
        return pd.DataFrame()

    higher_age_column = STANDARD_AGE_COLUMNS["HIGHER"]
    source_age = (
        higher_age_column
        if higher_age_column in ranking.columns
        else CANDLE_AGE_COLUMN
        if CANDLE_AGE_COLUMN in ranking.columns
        else None
    )
    if "Symbol" not in ranking.columns or source_age is None:
        return pd.DataFrame()

    source_columns = [
        "Symbol",
        source_age,
        REGIME_START_STANDARD_COLUMN,
        "Lower Regime",
        "Lower Bias",
        "Middle Regime",
        "Middle Bias",
        "Higher Regime",
        "Higher Bias",
        "Completed Candle",
        "Timeframe",
    ]
    table = ranking[[column for column in source_columns if column in ranking.columns]].copy()
    table = table.rename(columns={source_age: HIGHER_AGE_DISPLAY_COLUMN})
    table[HIGHER_AGE_DISPLAY_COLUMN] = pd.to_numeric(
        table[HIGHER_AGE_DISPLAY_COLUMN], errors="coerce"
    )
    table = (
        table.dropna(subset=[HIGHER_AGE_DISPLAY_COLUMN])
        .drop_duplicates(subset=["Symbol"], keep="first")
    )
    if table.empty:
        return table

    table[HIGHER_AGE_RANK_COLUMN] = (
        table[HIGHER_AGE_DISPLAY_COLUMN]
        .rank(method="min", ascending=True, na_option="bottom")
        .astype("Int64")
    )
    # The shared strategy resolver reads the canonical age alias.  Point it at
    # Higher's age for this table without changing the Middle table's value.
    table[CANDLE_AGE_COLUMN] = table[HIGHER_AGE_DISPLAY_COLUMN]
    # Preserve the old key for external exports that still expect it.
    table[AGE_ONLY_RANK_COLUMN] = table[HIGHER_AGE_RANK_COLUMN]

    table = table.rename(
        columns={
            "Higher Regime": "Higher Standard Regime State",
            "Higher Bias": "Higher Standard Regime Bias",
            "Middle Regime": "Middle Standard Regime State",
            "Middle Bias": "Middle Regime Bias",
            "Lower Regime": "Lower Standard Regime State",
            "Lower Bias": "Lower Regime Bias",
        }
    )
    # Use the canonical RAS/strategy rules so Higher's Open and Close views do
    # not silently disagree with the Middle table.
    table = _add_strategy_decision_column(table)

    ordered = [
        HIGHER_AGE_RANK_COLUMN,
        SL_DECISION_COLUMN,
        "Symbol",
        STRATEGY_DECISION_COLUMN,
        WEIGHTED_RISK_DECISION_COLUMN,
        "Lower Regime Bias",
        "Middle Regime Bias",
        "Higher Standard Regime Bias",
        HIGHER_AGE_DISPLAY_COLUMN,
        RAS_DECISION_COLUMN,
        "Higher Standard Regime State",
        "Middle Standard Regime State",
        "Lower Standard Regime State",
        REGIME_START_STANDARD_COLUMN,
        "Completed Candle",
        "Timeframe",
        AGE_ONLY_RANK_COLUMN,
    ]
    table = table[[column for column in ordered if column in table.columns]]
    return table.sort_values(
        [HIGHER_AGE_RANK_COLUMN, "Symbol"], kind="mergesort"
    ).reset_index(drop=True)


def _publish_higher_ranking_to_field_sections(
    state: MutableMapping[str, Any],
    higher_ranking: pd.DataFrame,
) -> None:
    """
    Publish Higher Standard Regime ranking to both Field Close and
    Field Open state locations.

    Field Close is explicitly designated as the first section.

    The actual Field Open / Field Close renderer in your application
    can consume these state objects without recomputing the ranking.
    """

    if not isinstance(
        higher_ranking,
        pd.DataFrame,
    ):
        higher_ranking = pd.DataFrame()

    ranking_copy = (
        higher_ranking.copy()
    )

    state[
        FIELD_CLOSE_HIGHER_RANKING_KEY
    ] = ranking_copy.copy()

    state[
        FIELD_OPEN_HIGHER_RANKING_KEY
    ] = ranking_copy.copy()

    # Mobile cards use the same canonical dataframe.
    state[
        FIELD_CLOSE_HIGHER_MOBILE_KEY
    ] = ranking_copy.copy()

    state[
        FIELD_OPEN_HIGHER_MOBILE_KEY
    ] = ranking_copy.copy()

    # Keep a single, auditable publication bundle.  Field Open/Close and the
    # standalone Field 3 renderer can therefore read the exact same Higher
    # frame instead of rebuilding a second version on demand.
    canonical = state.get(FIELD3_CANONICAL_TABLES_KEY)
    if not isinstance(canonical, MutableMapping):
        canonical = {}
    canonical["higher"] = ranking_copy.copy()
    state[FIELD3_CANONICAL_TABLES_KEY] = canonical

    # Explicit ordering.
    state[
        FIELD_SECTION_ORDER_KEY
    ] = (
        "FIELD_CLOSE",
        "FIELD_OPEN",
    )


def render_field_close_higher_ranking(
    st: Any,
    state: MutableMapping[str, Any],
) -> None:
    """Render the canonical Higher table in the Field Close form.

    Field Close is deliberately rendered first.  The dataframe is copied from
    the published state object, so opening this expander never recalculates or
    changes the Finder/Middle snapshot.
    """
    _render_higher_ranking_section(
        st,
        state,
        key=FIELD_CLOSE_HIGHER_RANKING_KEY,
        title="📈 Field Close — Higher Standard Regime Ranking (Open / Close)",
        key_suffix="close",
    )


def render_field_open_higher_ranking(
    st: Any,
    state: MutableMapping[str, Any],
) -> None:
    """Render the same canonical Higher table in the Field Open form."""
    _render_higher_ranking_section(
        st,
        state,
        key=FIELD_OPEN_HIGHER_RANKING_KEY,
        title="📈 Field Open — Higher Standard Regime Ranking (Open / Close)",
        key_suffix="open",
    )


def _render_higher_ranking_section(
    st: Any,
    state: MutableMapping[str, Any],
    *,
    key: str,
    title: str,
    key_suffix: str,
) -> None:
    """Render one of the two read-only Higher Standard Open/Close views."""
    higher = state.get(key)
    if not isinstance(higher, pd.DataFrame) or higher.empty:
        st.info(
            "Higher Standard Regime ranking is not available in the completed "
            "generation yet. Run Super Quick/Quick after loading symbols."
        )
        return

    with st.expander(title, expanded=False):
        st.caption(
            "This is the same published Higher Standard snapshot used by Field "
            "Open, Field Close, and the Middle/Finder decision rules. Opening or "
            "closing this section does not recalculate market data."
        )
        render_responsive_records(
            st,
            state,
            higher,
            preferred_columns=[
                HIGHER_AGE_RANK_COLUMN,
                "Symbol",
                STRATEGY_DECISION_COLUMN,
                "Lower Regime Bias",
                "Middle Regime Bias",
                "Higher Standard Regime Bias",
                HIGHER_AGE_DISPLAY_COLUMN,
                RAS_DECISION_COLUMN,
                "Higher Standard Regime State",
                "Middle Standard Regime State",
                "Lower Standard Regime State",
                "Completed Candle",
                "Timeframe",
            ],
            rank_column=HIGHER_AGE_RANK_COLUMN,
            desktop_height=min(620, 100 + 38 * len(higher)),
            full_table_label=f"Full Higher Standard Regime Ranking — Field {key_suffix.title()}",
            visualize_rank_bands=True,
            cell_extreme_highlights=[HIGHER_AGE_DISPLAY_COLUMN],
            highlight_min_cells=True,
            highlight_max_cells=True,
            force_table_on_phone=True,
            column_config={
                STRATEGY_DECISION_COLUMN: st.column_config.TextColumn(
                    STRATEGY_DECISION_COLUMN,
                    width="large",
                    help="Direction first, then strategy level: BUY/SELL — S1/S2/S3/S4/None.",
                ),
            },
        )


# ============================================================================
# Middle regime ranking
# ============================================================================

def _middle_age_ranking(
    ranking: pd.DataFrame,
) -> pd.DataFrame:

    if (
        not isinstance(
            ranking,
            pd.DataFrame,
        )
        or ranking.empty
    ):
        return pd.DataFrame()

    middle_age_column = (
        STANDARD_AGE_COLUMNS[
            "MIDDLE"
        ]
    )

    columns = [
        c
        for c in (
            "Symbol",
            middle_age_column,

            MIDDLE_AVG_SOURCE_COLUMN,
            MIDDLE_COMPLETED_SAMPLE_COLUMN,

            "Lower Regime",
            "Lower Bias",
            "Middle Regime",
            "Middle Bias",
            "Higher Regime",
            "Higher Bias",

            "Completed Candle",
            "Timeframe",
        )
        if c in ranking.columns
    ]

    if (
        "Symbol" not in columns
        or middle_age_column not in columns
    ):
        return pd.DataFrame()

    table = ranking[
        columns
    ].copy()

    # Normalize all calculation lanes and older saved generations to one
    # stable display column. Older rows remain visible with a blank value.
    starting_price_source = _first_existing_column(
        ranking,
        (
            STARTING_PRICE_COLUMN,
            "Middle Regime Starting Price",
            "Middle Starting Price",
            "Regime Start Price",
        ),
    )
    if starting_price_source is not None:
        table[STARTING_PRICE_COLUMN] = pd.to_numeric(
            ranking[starting_price_source],
            errors="coerce",
        ).round(5)
    else:
        table[STARTING_PRICE_COLUMN] = pd.Series(
            pd.NA,
            index=table.index,
            dtype="Float64",
        )

    # ----------------------------------------------------------------------
    # Starting age
    # ----------------------------------------------------------------------

    table[
        CANDLE_AGE_COLUMN
    ] = pd.to_numeric(
        table[
            middle_age_column
        ],
        errors="coerce",
    )

    # ----------------------------------------------------------------------
    # Average regime count
    # ----------------------------------------------------------------------

    if (
        MIDDLE_AVG_SOURCE_COLUMN
        in table.columns
    ):

        table[
            AVG_REGIME_CANDLE_COLUMN
        ] = pd.to_numeric(
            table[
                MIDDLE_AVG_SOURCE_COLUMN
            ],
            errors="coerce",
        ).round(2)

    if (
        MIDDLE_COMPLETED_SAMPLE_COLUMN
        in table.columns
    ):

        table[
            COMPLETED_SAMPLE_DISPLAY_COLUMN
        ] = pd.to_numeric(
            table[
                MIDDLE_COMPLETED_SAMPLE_COLUMN
            ],
            errors="coerce",
        ).astype("Int64")

    # ----------------------------------------------------------------------
    # Regime maturity
    # ----------------------------------------------------------------------

    if (
        AVG_REGIME_CANDLE_COLUMN
        in table.columns
        and
        COMPLETED_SAMPLE_DISPLAY_COLUMN
        in table.columns
    ):

        avg_candles = pd.to_numeric(
            table[
                AVG_REGIME_CANDLE_COLUMN
            ],
            errors="coerce",
        )

        sample_count = pd.to_numeric(
            table[
                COMPLETED_SAMPLE_DISPLAY_COLUMN
            ],
            errors="coerce",
        )

        table[
            CANDLE_STATISTICS_COLUMN
        ] = (
            table[
                CANDLE_AGE_COLUMN
            ]
            .div(
                avg_candles
            )
            .mul(
                np.log1p(
                    sample_count
                )
            )
            .where(
                (avg_candles > 0)
                & (sample_count > 0)
            )
            .round(2)
        )

    table = (
        table
        .dropna(
            subset=[
                CANDLE_AGE_COLUMN
            ]
        )
        .drop_duplicates(
            subset=[
                "Symbol"
            ],
            keep="first",
        )
    )

    if table.empty:
        return table

    # Establish the requested youngest-SA-first order before assigning a
    # strict rank; symbol is the deterministic tie-breaker for equal ages.
    table = table.sort_values(
        [CANDLE_AGE_COLUMN, "Symbol"],
        ascending=[True, True],
        kind="mergesort",
    ).reset_index(drop=True)

    table[
        MIDDLE_AGE_RANK_COLUMN
    ] = (
        table[
            CANDLE_AGE_COLUMN
        ]
        .rank(
            # A strict first-to-last priority is easier to read than tied
            # ranks: rank 1 is always the youngest SA in this snapshot.
            method="first",
            ascending=True,
        )
        .astype("Int64")
    )

    ordered = [
        MIDDLE_AGE_RANK_COLUMN,
        "Symbol",
        "Lower Bias",
        "Middle Bias",
        "Higher Bias",
        CANDLE_AGE_COLUMN,
        STARTING_PRICE_COLUMN,
        CANDLE_STATISTICS_COLUMN,
        AVG_REGIME_CANDLE_COLUMN,
        COMPLETED_SAMPLE_DISPLAY_COLUMN,
        "Middle Regime",
        "Lower Regime",
        "Higher Regime",
        "Completed Candle",
        "Timeframe",
    ]

    table = table[
        [
            c
            for c in ordered
            if c in table.columns
        ]
    ]

    table = table.rename(
        columns={
            "Middle Regime":
                "Middle Standard Regime State",
            "Middle Bias":
                "Middle Regime Bias",
            "Lower Regime":
                "Lower Standard Regime State",
            "Lower Bias":
                "Lower Regime Bias",
            "Higher Regime":
                "Higher Standard Regime State",
            "Higher Bias":
                "Higher Standard Regime Bias",
        }
    )

    # ----------------------------------------------------------------------
    # RAS
    # ----------------------------------------------------------------------

    table = _add_ras_decision_column(
        table
    )

    # ----------------------------------------------------------------------
    # Strategy Decision
    # ----------------------------------------------------------------------

    table = _add_strategy_decision_column(
        table
    )

    # ----------------------------------------------------------------------
    # FINAL COLUMN ORDER
    #
    # Symbol
    # Strategy Decision
    # Lower Bias
    # Middle Bias
    # Higher Bias
    # Starting Age
    # RAS
    #
    # RAS is immediately beside Starting Age.
    # ----------------------------------------------------------------------

    preferred_order = [
        MIDDLE_AGE_RANK_COLUMN,
        SL_DECISION_COLUMN,
        "Symbol",
        STRATEGY_DECISION_COLUMN,
        WEIGHTED_RISK_DECISION_COLUMN,
        WEIGHTED_RISK_SCORE_COLUMN,
        "Lower Regime Bias",
        "Middle Regime Bias",
        "Higher Standard Regime Bias",
        CANDLE_AGE_COLUMN,
        STARTING_PRICE_COLUMN,
        RISK_SCORE_COLUMN,
        IS_POTENTIAL_BLOCK_COLUMN,
        HOURLY_BLOCK_COUNT_COLUMN,
        FINAL_ACTION_COLUMN,
        "Hour",
        RAS_DECISION_COLUMN,
        CANDLE_STATISTICS_COLUMN,
        AVG_REGIME_CANDLE_COLUMN,
        COMPLETED_SAMPLE_DISPLAY_COLUMN,
        "Middle Standard Regime State",
        "Lower Standard Regime State",
        "Higher Standard Regime State",
        "Completed Candle",
        "Timeframe",
    ]

    table = table[
        [
            c
            for c in preferred_order
            if c in table.columns
        ]
    ]

    return (
        table
        .sort_values(
            [
                CANDLE_AGE_COLUMN,
                "Symbol",
            ],
            ascending=[
                True,
                True,
            ],
            kind="mergesort",
        )
        .reset_index(
            drop=True
        )
    )


# ============================================================================
# Datetime helpers
# ============================================================================

def _normalize_datetime_value(
    value: Any,
) -> pd.Timestamp | None:

    if value is None:
        return None

    try:
        parsed = pd.to_datetime(
            value,
            errors="coerce",
        )
    except Exception:
        return None

    if pd.isna(parsed):
        return None

    try:
        timestamp = pd.Timestamp(
            parsed
        )
    except Exception:
        return None

    try:

        if timestamp.tz is not None:
            timestamp = (
                timestamp.tz_localize(
                    None
                )
            )

    except Exception:

        try:
            timestamp = (
                timestamp.tz_convert(
                    None
                )
            )

        except Exception:
            pass

    return timestamp


def _normalize_datetime_series(
    values: Any,
) -> pd.Series:

    try:

        parsed = pd.to_datetime(
            values,
            errors="coerce",
        )

    except Exception:

        return pd.Series(
            index=getattr(
                values,
                "index",
                None,
            ),
            dtype="datetime64[ns]",
        )

    if isinstance(
        parsed,
        pd.DatetimeIndex,
    ):

        try:

            if parsed.tz is not None:
                parsed = parsed.tz_localize(
                    None
                )

        except Exception:
            pass

    elif isinstance(
        parsed,
        pd.Series,
    ):

        try:

            if (
                getattr(
                    parsed.dt,
                    "tz",
                    None,
                )
                is not None
            ):

                parsed = (
                    parsed.dt.tz_localize(
                        None
                    )
                )

        except Exception:

            parsed = parsed.apply(
                _normalize_datetime_value
            )

    return parsed


def _find_datetime_column(
    frame: pd.DataFrame,
) -> str | None:

    if (
        not isinstance(
            frame,
            pd.DataFrame,
        )
        or frame.empty
    ):
        return None

    candidates = (
        "Datetime",
        "DateTime",
        "datetime",
        "Timestamp",
        "timestamp",
        "Snapshot Datetime",
        "Snapshot DateTime",
        "Current Candle Datetime",
        "Current Candle DateTime",
        "Current Candle Timestamp",
        "Completed Candle Datetime",
        "Completed Candle DateTime",
        "Completed Candle Time",
        "Completed Candle",
        "Candle Datetime",
        "Candle DateTime",
        "Candle Time",
        "Time",
        "time",
        "Date",
        "date",
    )

    for column in candidates:

        if column not in frame.columns:
            continue

        values = frame[
            column
        ]

        if pd.api.types.is_bool_dtype(
            values
        ):
            continue

        parsed = (
            _normalize_datetime_series(
                values
            )
        )

        if (
            isinstance(
                parsed,
                pd.Series,
            )
            and parsed.notna().any()
        ):
            return column

    for column in frame.columns:

        name = str(
            column
        ).lower()

        if (
            "datetime" in name
            or "timestamp" in name
            or "candle time" in name
            or name == "date"
            or name == "time"
        ):

            values = frame[
                column
            ]

            if pd.api.types.is_bool_dtype(
                values
            ):
                continue

            parsed = (
                _normalize_datetime_series(
                    values
                )
            )

            if (
                isinstance(
                    parsed,
                    pd.Series,
                )
                and parsed.notna().any()
            ):
                return column

    return None


def _resolve_datetime_from_frame(
    frame: pd.DataFrame,
    candidates: tuple[str, ...],
) -> pd.Timestamp | None:

    if (
        not isinstance(
            frame,
            pd.DataFrame,
        )
        or frame.empty
    ):
        return None

    for column in candidates:

        if column not in frame.columns:
            continue

        values = frame[
            column
        ]

        if pd.api.types.is_bool_dtype(
            values
        ):
            continue

        parsed = (
            _normalize_datetime_series(
                values
            )
        )

        if (
            isinstance(
                parsed,
                pd.Series,
            )
            and parsed.notna().any()
        ):

            try:

                return pd.Timestamp(
                    parsed
                    .dropna()
                    .max()
                )

            except Exception:
                continue

    return None


def _resolve_current_candle_datetime(
    ranking: pd.DataFrame,
    state: MutableMapping[str, Any],
) -> pd.Timestamp | None:

    candidates = (
        "Snapshot Datetime",
        "Snapshot DateTime",
        "Snapshot Timestamp",
        "Current Candle Datetime",
        "Current Candle DateTime",
        "Current Candle Timestamp",
        "Candle Datetime",
        "Candle DateTime",
        "Candle Time",
        "Completed Candle Datetime",
        "Completed Candle DateTime",
        "Completed Candle Time",
        "Datetime",
        "DateTime",
        "datetime",
        "Timestamp",
        "timestamp",
        "Completed Candle",
    )

    resolved = (
        _resolve_datetime_from_frame(
            ranking,
            candidates,
        )
    )

    if resolved is not None:
        return resolved

    evidence = state.get(
        "field3_regime_evidence_v2"
    )

    resolved = (
        _resolve_datetime_from_frame(
            evidence,
            candidates,
        )
    )

    if resolved is not None:
        return resolved

    state_keys = (
        "field3_current_candle_datetime",
        "field3_current_candle_datetime_20260722",
        "field3_current_candle_timestamp",
        "current_candle_datetime",
        "current_candle_time",
        "current_candle_timestamp",
        "field3_candle_datetime",
        "field3_candle_time",
        "field3_candle_timestamp",
        "completed_candle_datetime",
        "completed_candle_time",
        "completed_candle_timestamp",
        "last_candle_datetime",
        "last_candle_time",
        "last_candle_timestamp",
        "last_completed_candle_datetime",
        "last_completed_candle_time",
        "current_datetime",
        "current_timestamp",
    )

    for key in state_keys:

        parsed = (
            _normalize_datetime_value(
                state.get(key)
            )
        )

        if parsed is not None:
            return parsed

    return None


# ============================================================================
# Historical storage
# ============================================================================

def _load_middle_regime_history(
    state: MutableMapping[str, Any],
) -> pd.DataFrame:

    frames: list[pd.DataFrame] = []

    try:

        if MIDDLE_HISTORY_FILE.exists():

            saved = pd.read_parquet(
                MIDDLE_HISTORY_FILE
            )

            if (
                isinstance(
                    saved,
                    pd.DataFrame,
                )
                and not saved.empty
            ):
                frames.append(
                    saved
                )

    except Exception as exc:

        state[
            "field3_middle_history_load_error_20260722"
        ] = (
            f"{type(exc).__name__}: {exc}"
        )

    try:
        if MIDDLE_HISTORY_CSV_FILE.exists():
            saved_csv = pd.read_csv(MIDDLE_HISTORY_CSV_FILE, low_memory=False)
            if isinstance(saved_csv, pd.DataFrame) and not saved_csv.empty:
                frames.append(saved_csv)
    except Exception as exc:
        state["field3_middle_history_csv_load_error_20260818"] = f"{type(exc).__name__}: {exc}"

    session_history = state.get(
        MIDDLE_HISTORY_STATE_KEY
    )

    if (
        isinstance(
            session_history,
            pd.DataFrame,
        )
        and not session_history.empty
    ):
        frames.append(
            session_history
        )

    if not frames:
        return pd.DataFrame()

    frames = _sanitize_frame_collection(frames)
    history = pd.concat(
        frames,
        ignore_index=True,
    )
    history = _sanitize_index_column_collision(history)

    if "Datetime" not in history.columns:

        fallback = _find_datetime_column(
            history
        )

        if fallback is None:
            return pd.DataFrame()

        history = history.rename(
            columns={
                fallback: "Datetime"
            }
        )

    history[
        "Datetime"
    ] = _normalize_datetime_series(
        history[
            "Datetime"
        ]
    )

    history = (
        history
        .dropna(
            subset=[
                "Datetime"
            ]
        )
        .drop_duplicates(
            subset=(
                [
                    "Datetime",
                    "Symbol",
                ]
                if "Symbol"
                in history.columns
                else [
                    "Datetime"
                ]
            ),
            keep="last",
        )
        .sort_values(
            (
                [
                    "Datetime",
                    "Symbol",
                ]
                if "Symbol"
                in history.columns
                else [
                    "Datetime"
                ]
            ),
            kind="mergesort",
        )
        .reset_index(
            drop=True
        )
    )

    # Backfill current additions and recompute the weighted MS/CC/SA score for
    # the complete persisted history after all sources have been merged.
    history = _add_ras_decision_column(history)
    if "Symbol" in history.columns:
        history = _add_weighted_risk_score(history, state)

    return history


def _save_middle_regime_history_snapshot(
    state: MutableMapping[str, Any],
    middle_age: pd.DataFrame,
    ranking: pd.DataFrame,
) -> None:

    if not isinstance(
        middle_age,
        pd.DataFrame,
    ):
        return

    if middle_age.empty:
        return

    current_datetime = (
        _resolve_current_candle_datetime(
            ranking,
            state,
        )
    )

    if current_datetime is None:

        state[
            "field3_middle_history_timestamp_error_20260722"
        ] = (
            "Could not determine the actual completed candle "
            "datetime. Middle Regime snapshot was NOT saved."
        )

        return

    current_datetime = (
        _normalize_datetime_value(
            current_datetime
        )
    )

    if current_datetime is None:
        return

    snapshot = _sanitize_index_column_collision(
        _add_weighted_risk_score(
            _sanitize_index_column_collision(middle_age),
            state,
        )
    )

    snapshot[
        "Datetime"
    ] = current_datetime

    history = (
        _load_middle_regime_history(
            state
        )
    )

    if (
        isinstance(
            history,
            pd.DataFrame,
        )
        and not history.empty
    ):

        history = pd.concat(
            [
                history,
                snapshot,
            ],
            ignore_index=True,
        )

    else:

        history = snapshot.copy()

    history[
        "Datetime"
    ] = _normalize_datetime_series(
        history[
            "Datetime"
        ]
    )

    history = history.dropna(
        subset=[
            "Datetime"
        ]
    )

    if "Symbol" in history.columns:

        history = history.drop_duplicates(
            subset=[
                "Datetime",
                "Symbol",
            ],
            keep="last",
        )

    history = (
        history
        .sort_values(
            (
                [
                    "Datetime",
                    "Symbol",
                ]
                if "Symbol"
                in history.columns
                else [
                    "Datetime"
                ]
            ),
            kind="mergesort",
        )
        .reset_index(
            drop=True
        )
    )

    state[
        MIDDLE_HISTORY_STATE_KEY
    ] = history

    try:

        MIDDLE_HISTORY_FILE.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        history.to_parquet(
            MIDDLE_HISTORY_FILE,
            index=False,
        )
        history.to_csv(
            MIDDLE_HISTORY_CSV_FILE,
            index=False,
            encoding="utf-8",
        )

        state[
            "field3_middle_history_last_saved_20260722"
        ] = str(
            current_datetime
        )

        state[
            "field3_middle_history_count_20260722"
        ] = len(
            history
        )

    except Exception as exc:

        state[
            "field3_middle_history_save_error_20260722"
        ] = (
            f"{type(exc).__name__}: {exc}"
        )


# ============================================================================
# Finder helpers
# ============================================================================

def _make_finder_datetime(
    day: Any,
    clock_time: Any,
) -> pd.Timestamp | None:
    """Build an exact Finder datetime from free date/time inputs."""
    try:
        date_part = pd.Timestamp(day).normalize()
        parsed = pd.to_datetime(str(clock_time), errors="coerce")
        if pd.isna(parsed):
            return None
        return date_part + pd.Timedelta(
            hours=int(parsed.hour),
            minutes=int(parsed.minute),
            seconds=int(parsed.second),
        )
    except Exception:
        return None


def _weekday_dates(
    history: pd.DataFrame,
) -> list[Any]:

    if (
        not isinstance(
            history,
            pd.DataFrame,
        )
        or history.empty
        or "_FinderDatetime"
        not in history.columns
    ):
        return []

    values = (
        _normalize_datetime_series(
            history[
                "_FinderDatetime"
            ]
        )
    )

    values = values.loc[
        values.dt.weekday < 5
    ]

    return sorted(
        values
        .dt.date
        .dropna()
        .unique()
        .tolist()
    )


def _prepare_middle_history(
    history: pd.DataFrame,
) -> pd.DataFrame:

    if (
        not isinstance(
            history,
            pd.DataFrame,
        )
        or history.empty
    ):
        return pd.DataFrame()

    table = _sanitize_index_column_collision(history)

    datetime_column = _find_datetime_column(
        table
    )

    if datetime_column is None:
        return pd.DataFrame()

    table[
        "_FinderDatetime"
    ] = _normalize_datetime_series(
        table[
            datetime_column
        ]
    )

    table = table.dropna(
        subset=[
            "_FinderDatetime"
        ]
    ).copy()

    table = _add_weighted_risk_score(table)

    return (
        table
        .sort_values(
            "_FinderDatetime",
            kind="mergesort",
        )
        .reset_index(
            drop=True
        )
    )


def _finder_range(
    history: pd.DataFrame,
    start_dt: pd.Timestamp,
    end_dt: pd.Timestamp,
) -> pd.DataFrame:

    if not isinstance(history, pd.DataFrame) or history.empty:
        return pd.DataFrame()

    if "_FinderDatetime" not in history.columns:
        history = _prepare_middle_history(history)
    if history.empty or "_FinderDatetime" not in history.columns:
        return pd.DataFrame()

    start_value = _normalize_datetime_value(start_dt)
    end_value = _normalize_datetime_value(end_dt)
    if start_value is None or end_value is None or start_value > end_value:
        return pd.DataFrame()

    timestamps = (
        _normalize_datetime_series(
            history[
                "_FinderDatetime"
            ]
        )
    )

    mask = (
        (timestamps >= start_value)
        & (timestamps <= end_value)
    )

    selected = history.loc[mask].copy()
    sort_columns = ["_FinderDatetime"]
    if CANDLE_AGE_COLUMN in selected.columns:
        # Within each snapshot, keep the same youngest-SA-first priority as
        # the live Middle Standard table.
        selected[CANDLE_AGE_COLUMN] = pd.to_numeric(selected[CANDLE_AGE_COLUMN], errors="coerce")
        sort_columns.append(CANDLE_AGE_COLUMN)
    if "Symbol" in selected.columns:
        sort_columns.append("Symbol")
    return selected.sort_values(sort_columns, kind="mergesort").reset_index(drop=True)


def _filter_finder_result(
    data: pd.DataFrame,
    *,
    symbols: list[str] | tuple[str, ...] | None = None,
    strategies: list[str] | tuple[str, ...] | None = None,
) -> pd.DataFrame:
    """Apply optional preview/export filters without changing the full result."""
    if not isinstance(data, pd.DataFrame) or data.empty:
        return pd.DataFrame()
    filtered = data.copy()
    normalized_symbols = {str(value).strip().upper() for value in (symbols or []) if str(value).strip()}
    if normalized_symbols and "Symbol" in filtered.columns:
        filtered = filtered.loc[
            filtered["Symbol"].astype(str).str.upper().isin(normalized_symbols)
        ].copy()
    normalized_strategies = {str(value).strip() for value in (strategies or []) if str(value).strip()}
    if normalized_strategies and STRATEGY_DECISION_COLUMN in filtered.columns:
        filtered = filtered.loc[
            filtered[STRATEGY_DECISION_COLUMN].astype(str).isin(normalized_strategies)
        ].copy()
    return filtered.reset_index(drop=True)


def _finder_page(
    data: pd.DataFrame,
    *,
    page: int,
    page_size: int,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """Return one bounded preview page while full-range downloads stay complete."""
    total_rows = int(len(data)) if isinstance(data, pd.DataFrame) else 0
    size = max(10, min(int(page_size or 50), 500))
    total_pages = max(1, int(np.ceil(total_rows / size)))
    page_number = max(1, min(int(page or 1), total_pages))
    start = (page_number - 1) * size
    end = min(total_rows, start + size)
    frame = data.iloc[start:end].copy() if total_rows else pd.DataFrame(columns=getattr(data, "columns", []))
    return frame, {
        "page": page_number,
        "page_size": size,
        "total_pages": total_pages,
        "total_rows": total_rows,
        "start_row": start + 1 if total_rows else 0,
        "end_row": end,
    }


def _time_slot_table(
    data: pd.DataFrame,
    slot: str,
) -> pd.DataFrame:

    if (
        not isinstance(
            data,
            pd.DataFrame,
        )
        or data.empty
    ):
        return pd.DataFrame()

    timestamps = (
        _normalize_datetime_series(
            data[
                "_FinderDatetime"
            ]
        )
    )

    hour = int(
        slot.split(":")[0]
    )

    minute = 0

    mask = (
        (timestamps.dt.hour == hour)
        & (timestamps.dt.minute == minute)
        & (
            timestamps.dt.weekday < 5
        )
    )

    return data.loc[
        mask
    ].copy()


def _finder_csv(
    data: pd.DataFrame,
    start_dt: pd.Timestamp,
    end_dt: pd.Timestamp,
) -> str:
    """Export the entire selected Finder range as one flat CSV table."""
    if not isinstance(data, pd.DataFrame):
        return ""

    selected = _sanitize_index_column_collision(data.copy())
    if "_FinderDatetime" in selected.columns:
        selected["Datetime"] = _normalize_datetime_series(
            selected["_FinderDatetime"]
        )
        selected = selected.drop(
            columns=["_FinderDatetime"],
            errors="ignore",
        )

    if "Datetime" in selected.columns:
        dt = _normalize_datetime_series(selected["Datetime"])
        selected["Datetime"] = dt.dt.strftime("%Y-%m-%d %H:%M:%S")
    selected = _canonical_middle_export_frame(selected, datetime_first=True)

    if "Datetime" in selected.columns:
        sort_columns = ["Datetime"]
        if "Symbol" in selected.columns:
            sort_columns.append("Symbol")
        selected = selected.sort_values(
            sort_columns,
            kind="mergesort",
        ).reset_index(drop=True)

    return selected.to_csv(
        index=False,
        encoding="utf-8",
    )


# ============================================================================
# Excel styling
# ============================================================================

def _style_excel_sheet(
    worksheet: Any,
    dataframe: pd.DataFrame,
) -> None:

    if (
        dataframe.empty
        or PatternFill is None
    ):
        return

    fill_header = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    fill_min = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    fill_max = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    fill_match = PatternFill(
        fill_type="solid",
        fgColor="DDEBF7",
    )

    fill_top = PatternFill(
        fill_type="solid",
        fgColor="E2F0D9",
    )

    fill_bottom = PatternFill(
        fill_type="solid",
        fgColor="FCE4D6",
    )

    fill_allow = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    fill_disallow = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    fill_s1 = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    fill_s2 = PatternFill(
        fill_type="solid",
        fgColor="FCE4D6",
    )

    fill_s3 = PatternFill(
        fill_type="solid",
        fgColor="E4DFEC",
    )

    fill_none = PatternFill(
        fill_type="solid",
        fgColor="F2F2F2",
    )

    for cell in worksheet[1]:

        cell.fill = fill_header

        if Font is not None:
            cell.font = Font(
                bold=True
            )

        if Alignment is not None:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    # ----------------------------------------------------------------------
    # Numeric extremes
    # ----------------------------------------------------------------------

    extremes: dict[
        str,
        tuple[Any, Any],
    ] = {}

    for column in (
        MIDDLE_NUMERIC_HIGHLIGHT_COLUMNS
    ):

        if column not in dataframe.columns:
            continue

        values = pd.to_numeric(
            dataframe[column],
            errors="coerce",
        ).dropna()

        if values.empty:
            continue

        extremes[column] = (
            values.min(),
            values.max(),
        )

    # ----------------------------------------------------------------------
    # Matching bias rows
    # ----------------------------------------------------------------------

    matching_indexes: set[Any] = set()

    if (
        "Middle Regime Bias"
        in dataframe.columns
        and
        "Higher Standard Regime Bias"
        in dataframe.columns
    ):

        middle_bias = (
            dataframe[
                "Middle Regime Bias"
            ]
            .astype(str)
            .str.strip()
            .str.upper()
        )

        higher_bias = (
            dataframe[
                "Higher Standard Regime Bias"
            ]
            .astype(str)
            .str.strip()
            .str.upper()
        )

        matching_indexes = set(
            dataframe.index[
                middle_bias.eq(
                    higher_bias
                )
            ]
        )

    # ----------------------------------------------------------------------
    # Top / bottom
    # ----------------------------------------------------------------------

    top_indexes: set[Any] = set()
    bottom_indexes: set[Any] = set()

    if (
        MIDDLE_AGE_RANK_COLUMN
        in dataframe.columns
    ):

        ranks = pd.to_numeric(
            dataframe[
                MIDDLE_AGE_RANK_COLUMN
            ],
            errors="coerce",
        ).dropna()

        if not ranks.empty:

            ordered = (
                ranks
                .sort_values()
                .index
                .tolist()
            )

            top_indexes = set(
                ordered[:5]
            )

            bottom_indexes = set(
                ordered[-5:]
            )

    # ----------------------------------------------------------------------
    # Rows
    # ----------------------------------------------------------------------

    for excel_row, df_index in enumerate(
        dataframe.index,
        start=2,
    ):

        if (
            df_index
            in matching_indexes
        ):
            row_fill = fill_match

        elif (
            df_index
            in top_indexes
        ):
            row_fill = fill_top

        elif (
            df_index
            in bottom_indexes
        ):
            row_fill = fill_bottom

        else:
            row_fill = None

        if row_fill is not None:

            for cell in worksheet[
                excel_row
            ]:

                cell.fill = row_fill

        # --------------------------------------------------------------
        # Numeric min/max
        # --------------------------------------------------------------

        for column, (
            minimum,
            maximum,
        ) in extremes.items():

            column_index = (
                dataframe.columns.get_loc(
                    column
                )
                + 1
            )

            value = pd.to_numeric(
                dataframe.loc[
                    df_index,
                    column,
                ],
                errors="coerce",
            )

            if pd.isna(value):
                continue

            cell = worksheet.cell(
                row=excel_row,
                column=column_index,
            )

            if (
                value == minimum
                and value == maximum
            ):
                cell.fill = fill_min

            elif value == minimum:
                cell.fill = fill_min

            elif value == maximum:
                cell.fill = fill_max

        # --------------------------------------------------------------
        # RAS
        # --------------------------------------------------------------

        if (
            RAS_DECISION_COLUMN
            in dataframe.columns
        ):

            ras_index = (
                dataframe.columns.get_loc(
                    RAS_DECISION_COLUMN
                )
                + 1
            )

            ras_cell = worksheet.cell(
                row=excel_row,
                column=ras_index,
            )

            ras_value = str(
                dataframe.loc[
                    df_index,
                    RAS_DECISION_COLUMN,
                ]
            ).upper()

            if ras_value == "ALLOW":
                ras_cell.fill = fill_allow

            else:
                ras_cell.fill = fill_disallow

        # --------------------------------------------------------------
        # Weighted risk decision
        # --------------------------------------------------------------

        if WEIGHTED_RISK_DECISION_COLUMN in dataframe.columns:
            weighted_index = dataframe.columns.get_loc(WEIGHTED_RISK_DECISION_COLUMN) + 1
            weighted_cell = worksheet.cell(row=excel_row, column=weighted_index)
            weighted_value = str(dataframe.loc[df_index, WEIGHTED_RISK_DECISION_COLUMN]).upper()
            weighted_cell.fill = fill_disallow if weighted_value == "BLOCK" else fill_allow

        if FINAL_ACTION_COLUMN in dataframe.columns:
            action_index = dataframe.columns.get_loc(FINAL_ACTION_COLUMN) + 1
            action_cell = worksheet.cell(row=excel_row, column=action_index)
            action_value = str(dataframe.loc[df_index, FINAL_ACTION_COLUMN]).upper()
            action_cell.fill = fill_disallow if action_value == "BLOCK" else fill_allow

        # --------------------------------------------------------------
        # Strategy
        # --------------------------------------------------------------

        if (
            STRATEGY_DECISION_COLUMN
            in dataframe.columns
        ):

            strategy_index = (
                dataframe.columns.get_loc(
                    STRATEGY_DECISION_COLUMN
                )
                + 1
            )

            strategy_cell = worksheet.cell(
                row=excel_row,
                column=strategy_index,
            )

            strategy = str(
                dataframe.loc[
                    df_index,
                    STRATEGY_DECISION_COLUMN,
                ]
            ).upper()

            if strategy.startswith("S1"):
                strategy_cell.fill = fill_s1

            elif strategy.startswith("S2"):
                strategy_cell.fill = fill_s2

            elif strategy.startswith("S3"):
                strategy_cell.fill = fill_s3

            elif strategy.startswith("S4"):
                strategy_cell.fill = fill_disallow

            else:
                strategy_cell.fill = fill_none

    # ----------------------------------------------------------------------
    # Width
    # ----------------------------------------------------------------------

    if (
        get_column_letter is not None
    ):

        for column_index, column_name in enumerate(
            dataframe.columns,
            start=1,
        ):

            maximum = len(
                str(column_name)
            )

            for cell in worksheet[
                get_column_letter(
                    column_index
                )
            ]:

                if cell.value is not None:

                    maximum = max(
                        maximum,
                        len(
                            str(
                                cell.value
                            )
                        ),
                    )

            worksheet.column_dimensions[
                get_column_letter(
                    column_index
                )
            ].width = min(
                max(
                    maximum + 2,
                    12,
                ),
                42,
            )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )


def _dataframe_to_excel(
    dataframe: pd.DataFrame,
    sheet_name: str,
) -> bytes | None:

    if (
        Workbook is None
        or not isinstance(
            dataframe,
            pd.DataFrame,
        )
    ):
        return None

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = (
        sheet_name[:31]
    )

    # Header
    for column_index, column_name in enumerate(
        dataframe.columns,
        start=1,
    ):

        worksheet.cell(
            row=1,
            column=column_index,
            value=str(
                column_name
            ),
        )

    # Data.  Excel/openpyxl does not accept timezone-aware datetime objects;
    # normalize every datetime-like value to a timezone-naive Python value (or
    # a formatted string) before assigning cells.
    for row_index, (_, row) in enumerate(
        dataframe.iterrows(),
        start=2,
    ):

        for column_index, value in enumerate(
            row,
            start=1,
        ):

            if isinstance(value, (pd.Timestamp,)):
                value = _normalize_datetime_value(value)
                if value is not None:
                    value = value.to_pydatetime().replace(tzinfo=None)
            elif isinstance(value, (np.datetime64,)):
                parsed = _normalize_datetime_value(value)
                value = parsed.to_pydatetime().replace(tzinfo=None) if parsed is not None else None
            elif getattr(value, "tzinfo", None) is not None:
                value = value.replace(tzinfo=None)
            try:
                if pd.isna(value):
                    value = None
            except Exception:
                pass

            if isinstance(value, (list, tuple, dict, set)):
                value = str(value)

            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    _style_excel_sheet(
        worksheet,
        dataframe,
    )

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output.getvalue()


def _canonical_middle_export_frame(
    dataframe: pd.DataFrame,
    state: Mapping[str, Any] | None = None,
    *,
    datetime_first: bool = False,
) -> pd.DataFrame:
    """Return the exact Field 3 Middle layout used by the visible table.

    Risk/SL columns are added after the ranking builder, so relying on the
    dataframe's insertion order used to append SL near the end of CSV/XLSX
    exports.  Keep one explicit order for the UI and both download formats;
    unknown diagnostic columns are retained after the canonical columns.
    """
    if not isinstance(dataframe, pd.DataFrame):
        return pd.DataFrame()
    result = _sanitize_index_column_collision(dataframe.copy())
    result = _add_weighted_risk_score(result, state)
    preferred = [
        MIDDLE_AGE_RANK_COLUMN,
        SL_DECISION_COLUMN,
        "Symbol",
        STRATEGY_DECISION_COLUMN,
        WEIGHTED_RISK_DECISION_COLUMN,
        WEIGHTED_RISK_SCORE_COLUMN,
        "Lower Regime Bias",
        "Middle Regime Bias",
        "Higher Standard Regime Bias",
        CANDLE_AGE_COLUMN,
        STARTING_PRICE_COLUMN,
        RISK_SCORE_COLUMN,
        IS_POTENTIAL_BLOCK_COLUMN,
        HOURLY_BLOCK_COUNT_COLUMN,
        FINAL_ACTION_COLUMN,
        "Hour",
        RAS_DECISION_COLUMN,
        CANDLE_STATISTICS_COLUMN,
        AVG_REGIME_CANDLE_COLUMN,
        COMPLETED_SAMPLE_DISPLAY_COLUMN,
        "Middle Standard Regime State",
        "Lower Standard Regime State",
        "Higher Standard Regime State",
        "Completed Candle",
        "Timeframe",
    ]
    if datetime_first and "Datetime" in result.columns:
        preferred = ["Datetime"] + preferred
    ordered = [column for column in preferred if column in result.columns]
    remaining = [column for column in result.columns if column not in ordered]
    return result[ordered + remaining].copy()


# ============================================================================
# Finder Excel
# ============================================================================

def _finder_to_excel(
    data: pd.DataFrame,
    start_dt: pd.Timestamp,
    end_dt: pd.Timestamp,
) -> bytes | None:

    if (
        Workbook is None
        or not isinstance(data, pd.DataFrame)
        or data.empty
        or len(data) > 1_048_575
    ):
        return None

    workbook = Workbook()

    # A single flat worksheet scales to long historical ranges.  The former
    # one-sheet-per-snapshot layout could create thousands of worksheets,
    # duplicate worksheet names across months, and freeze the Streamlit run.
    export = _sanitize_index_column_collision(data.copy())
    if "_FinderDatetime" in export.columns:
        timestamps = _normalize_datetime_series(export["_FinderDatetime"])
        # ``Datetime`` may already be present in history loaded from CSV or
        # Parquet.  Replace it instead of calling DataFrame.insert on an
        # existing label.
        export = export.drop(columns=["Datetime"], errors="ignore")
        export.insert(
            0,
            "Datetime",
            timestamps.dt.strftime("%Y-%m-%d %H:%M:%S"),
        )
        export = export.drop(columns=["_FinderDatetime"], errors="ignore")
    export = _canonical_middle_export_frame(export, datetime_first=True)
    sort_columns = [column for column in ["Datetime", "Symbol"] if column in export.columns]
    if sort_columns:
        export = export.sort_values(sort_columns, kind="mergesort").reset_index(drop=True)

    worksheet = workbook.active
    worksheet.title = "Finder Results"

    for column_index, column_name in enumerate(export.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=str(column_name))

    for row_index, row in enumerate(export.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row, start=1):
            if isinstance(value, (pd.Timestamp,)):
                normalized = _normalize_datetime_value(value)
                value = normalized.to_pydatetime().replace(tzinfo=None) if normalized is not None else None
            elif isinstance(value, (np.datetime64,)):
                normalized = _normalize_datetime_value(value)
                value = normalized.to_pydatetime().replace(tzinfo=None) if normalized is not None else None
            elif getattr(value, "tzinfo", None) is not None:
                value = value.replace(tzinfo=None)
            try:
                if pd.isna(value):
                    value = None
            except Exception:
                pass
            if isinstance(value, (list, tuple, dict, set)):
                value = str(value)
            worksheet.cell(row=row_index, column=column_index, value=value)

    _style_excel_sheet(worksheet, export)

    # Range information
    info = workbook.create_sheet(
        title="Range"
    )

    info["A1"] = "Start"
    info["B1"] = start_dt.strftime(
        "%Y-%m-%d %H:%M"
    )

    info["A2"] = "End"
    info["B2"] = end_dt.strftime(
        "%Y-%m-%d %H:%M"
    )

    info["A3"] = "Calendar Days"
    info["B3"] = "Any day selectable"

    info["A4"] = "Time Selection"
    info["B4"] = "Any time selectable"

    info["A5"] = "Rows"
    info["B5"] = len(export)

    info["A6"] = "Layout"
    info["B6"] = "One scalable flat Finder Results worksheet"

    info.column_dimensions["A"].width = 20
    info.column_dimensions["B"].width = 48

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output.getvalue()


# ============================================================================
# Independent Finder
# ============================================================================

def render_middle_regime_finder(
    st: Any,
    state: MutableMapping[str, Any],
) -> None:
    """Independent Finder with fully unlocked date and time controls."""

    st.caption(
        "Choose any calendar date and any time for both endpoints. "
        "The range is inclusive and searches every stored historical record."
    )

    history = _prepare_middle_history(
        _load_middle_regime_history(state)
    )
    if history.empty:
        st.info("No historical Middle Regime records are available yet.")
        return

    history_times = _normalize_datetime_series(
        history["_FinderDatetime"]
    ).dropna()
    min_day = (
        history_times.min().date()
        if not history_times.empty
        else pd.Timestamp.today().date()
    )
    max_day = (
        history_times.max().date()
        if not history_times.empty
        else pd.Timestamp.today().date()
    )

    if not history_times.empty:
        st.caption(
            "Loaded Finder coverage: "
            f"{history_times.min().strftime('%Y-%m-%d %H:%M')} → "
            f"{history_times.max().strftime('%Y-%m-%d %H:%M')} · "
            f"{len(history):,} historical row(s)."
        )

    range_mode = st.selectbox(
        "Finder Range",
        ["Custom", "Last 7 Days", "Last 30 Days", "Last 365 Days", "Last 2 Years", "All Loaded History"],
        index=0,
        key="middle_finder_range_mode_20260818",
        help="Presets use the latest loaded historical candle. Custom keeps exact date/time selection.",
    )

    try:
        default_start_day = pd.Timestamp(
            state.get(FINDER_START_DAY_STATE_KEY, min_day)
        ).date()
    except Exception:
        default_start_day = min_day

    try:
        default_end_day = pd.Timestamp(
            state.get(FINDER_END_DAY_STATE_KEY, max_day)
        ).date()
    except Exception:
        default_end_day = max_day

    if not history_times.empty and range_mode != "Custom":
        latest = pd.Timestamp(history_times.max())
        earliest = pd.Timestamp(history_times.min())
        # Presets are calendar-relative to today, not to the stale first row
        # of an older saved generation.  This prevents "Last 365 Days" from
        # silently becoming 2026-02-01 → 2026-08-24 when the requested window
        # is 2025-08-24 → 2026-08-24.
        preset_end = max(latest, pd.Timestamp.today().normalize())
        preset_days = {
            "Last 7 Days": 7,
            "Last 30 Days": 30,
            "Last 365 Days": 365,
            "Last 2 Years": 730,
        }
        if range_mode == "All Loaded History":
            preset_start = earliest
        else:
            preset_start = preset_end - pd.Timedelta(days=preset_days[range_mode])
        default_start_day = preset_start.date()
        default_end_day = (latest if range_mode == "All Loaded History" else preset_end).date()
        state[FINDER_START_TIME_STATE_KEY] = preset_start.strftime("%H:%M:%S")
        state[FINDER_END_TIME_STATE_KEY] = (latest if range_mode == "All Loaded History" else preset_end).strftime("%H:%M:%S")

    def _safe_time(value: Any, fallback: str) -> Any:
        try:
            parsed = pd.to_datetime(str(value), errors="coerce")
            if pd.isna(parsed):
                raise ValueError
            return parsed.time()
        except Exception:
            return pd.Timestamp(fallback).time()

    start_col, end_col = st.columns(2)

    with start_col:
        st.markdown("##### Starting Point")
        start_day = st.date_input(
            "Starting Date",
            value=default_start_day,
            key="middle_finder_start_date_widget",
            disabled=(range_mode != "Custom"),
        )
        start_time = st.time_input(
            "Starting Time",
            value=_safe_time(
                state.get(FINDER_START_TIME_STATE_KEY, "00:00"),
                "00:00",
            ),
            step=60,
            key="middle_finder_start_time_widget",
            disabled=(range_mode != "Custom"),
        )

    with end_col:
        st.markdown("##### Ending Point")
        end_day = st.date_input(
            "Ending Date",
            value=default_end_day,
            key="middle_finder_end_date_widget",
            disabled=(range_mode != "Custom"),
        )
        end_time = st.time_input(
            "Ending Time",
            value=_safe_time(
                state.get(FINDER_END_TIME_STATE_KEY, "23:59"),
                "23:59",
            ),
            step=60,
            key="middle_finder_end_time_widget",
            disabled=(range_mode != "Custom"),
        )

    state[FINDER_START_DAY_STATE_KEY] = start_day
    state[FINDER_START_TIME_STATE_KEY] = start_time.strftime("%H:%M:%S")
    state[FINDER_END_DAY_STATE_KEY] = end_day
    state[FINDER_END_TIME_STATE_KEY] = end_time.strftime("%H:%M:%S")

    start_dt = _make_finder_datetime(start_day, start_time)
    end_dt = _make_finder_datetime(end_day, end_time)

    # Presets are calculated from actual loaded history, not stale widget state.
    if not history_times.empty and range_mode != "Custom":
        latest = pd.Timestamp(history_times.max())
        earliest = pd.Timestamp(history_times.min())
        preset_end = max(latest, pd.Timestamp.today().normalize())
        preset_days = {"Last 7 Days": 7, "Last 30 Days": 30, "Last 365 Days": 365, "Last 2 Years": 730}
        if range_mode == "All Loaded History":
            start_dt = earliest
            end_dt = latest
        else:
            start_dt = preset_end - pd.Timedelta(days=preset_days[range_mode])
            end_dt = preset_end

    if start_dt is None or end_dt is None:
        st.error("The selected Finder date/time is invalid.")
        return
    if start_dt > end_dt:
        st.error("Ending Date/Time must be after Starting Date/Time.")
        return

    st.caption(
        "Date and time are not locked to available historical days or fixed time slots."
    )

    result = _finder_range(history, start_dt, end_dt)
    state[FINDER_RESULT_STATE_KEY] = result.copy()
    state[FINDER_LAST_RANGE_STATE_KEY] = {
        "start": start_dt,
        "end": end_dt,
    }

    if result.empty:
        coverage_start = history_times.min().strftime("%Y-%m-%d %H:%M") if not history_times.empty else "—"
        coverage_end = history_times.max().strftime("%Y-%m-%d %H:%M") if not history_times.empty else "—"
        st.warning(
            "No Finder rows exist inside this exact range. "
            f"Current loaded historical coverage is {coverage_start} → {coverage_end}. "
            "Run Load & Update + Super Quick to rebuild Finder history from the loaded candles, "
            "then choose a range inside that coverage."
        )
        st.caption("No empty/header-only CSV is generated for a zero-row range.")
        return

    timestamps = _normalize_datetime_series(result["_FinderDatetime"])
    available_snapshots = sorted(
        timestamps.dropna().unique().tolist()
    )

    st.success(
        "Finder range loaded: "
        f"{start_dt.strftime('%Y-%m-%d %H:%M')} → "
        f"{end_dt.strftime('%Y-%m-%d %H:%M')}"
    )
    st.caption(
        f"{len(available_snapshots):,} snapshot time(s) found · "
        f"{len(result):,} row(s) · endpoints included"
    )

    finder_name = (
        "middle_regime_finder_"
        f"{start_dt.strftime('%Y%m%d_%H%M')}_to_"
        f"{end_dt.strftime('%Y%m%d_%H%M')}"
    )
    finder_csv = _finder_csv(result, start_dt, end_dt)
    st.download_button(
        label="📥 Download Full Finder CSV",
        data=finder_csv,
        file_name=f"{finder_name}.csv",
        mime="text/csv",
        key="middle_finder_download_csv",
        use_container_width=True,
        help="Always contains every row in the inclusive selected range.",
    )

    filter_left, filter_right = st.columns(2)
    symbol_options = (
        sorted(result["Symbol"].dropna().astype(str).unique().tolist())
        if "Symbol" in result.columns
        else []
    )
    strategy_options = (
        sorted(result[STRATEGY_DECISION_COLUMN].dropna().astype(str).unique().tolist())
        if STRATEGY_DECISION_COLUMN in result.columns
        else []
    )
    with filter_left:
        selected_symbols = st.multiselect(
            "Filter Symbols",
            symbol_options,
            default=[],
            key="middle_finder_symbol_filter_20260818",
            help="Leave empty to include every symbol.",
        )
    with filter_right:
        selected_strategies = st.multiselect(
            "Filter Strategies",
            strategy_options,
            default=[],
            key="middle_finder_strategy_filter_20260818",
            help="Leave empty to include every strategy.",
        )

    filtered = _filter_finder_result(
        result,
        symbols=selected_symbols,
        strategies=selected_strategies,
    )
    st.caption(f"Preview filters: {len(filtered):,} of {len(result):,} row(s).")
    if filtered.empty:
        st.warning("The current symbol/strategy filters match no Finder rows.")
        return

    if len(filtered) != len(result):
        st.download_button(
            label="📥 Download Filtered Finder CSV",
            data=_finder_csv(filtered, start_dt, end_dt),
            file_name=f"{finder_name}_filtered.csv",
            mime="text/csv",
            key="middle_finder_download_filtered_csv_20260818",
            use_container_width=True,
        )

    if len(filtered) <= 1_048_575:
        prepare_excel = st.toggle(
            "Prepare colored Excel (optional)",
            value=False,
            key="middle_finder_prepare_excel_20260818",
            help="CSV is immediate. Enable this only when you need a styled XLSX file.",
        )
        if prepare_excel:
            with st.spinner("Preparing one scalable Finder Results worksheet…"):
                finder_excel = _finder_to_excel(filtered, start_dt, end_dt)
            if finder_excel is not None:
                st.download_button(
                    label="🎨 Download Colored Excel",
                    data=finder_excel,
                    file_name=f"{finder_name}.xlsx",
                    mime=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                    key="middle_finder_download_excel",
                    use_container_width=True,
                )
            else:
                st.warning("Excel support is unavailable. Use the complete CSV download above.")
    else:
        st.info("This result exceeds Excel's row limit; use the complete CSV download.")

    preferred_columns = [
        "Datetime",
        MIDDLE_AGE_RANK_COLUMN,
        SL_DECISION_COLUMN,
        "Symbol",
        STRATEGY_DECISION_COLUMN,
        "Lower Regime Bias",
        "Middle Regime Bias",
        "Higher Standard Regime Bias",
        CANDLE_AGE_COLUMN,
        STARTING_PRICE_COLUMN,
        RISK_SCORE_COLUMN,
        IS_POTENTIAL_BLOCK_COLUMN,
        HOURLY_BLOCK_COUNT_COLUMN,
        FINAL_ACTION_COLUMN,
        "Hour",
        WEIGHTED_RISK_SCORE_COLUMN,
        RAS_DECISION_COLUMN,
        CANDLE_STATISTICS_COLUMN,
        AVG_REGIME_CANDLE_COLUMN,
        COMPLETED_SAMPLE_DISPLAY_COLUMN,
        "Middle Standard Regime State",
        "Lower Standard Regime State",
        "Higher Standard Regime State",
    ]
    view_mode = st.selectbox(
        "Finder Preview",
        ["Flat table (paged)", "Single snapshot"],
        key=FINDER_VIEW_MODE_STATE_KEY,
        help="The preview is bounded so changing pages never renders the entire history at once.",
    )

    filtered_times = _normalize_datetime_series(filtered["_FinderDatetime"])
    if view_mode == "Single snapshot":
        snapshot_options = sorted(filtered_times.dropna().unique().tolist(), reverse=True)
        snapshot_key = "middle_finder_snapshot_picker_20260818"
        if state.get(snapshot_key) not in snapshot_options:
            state.pop(snapshot_key, None)
        raw_timestamp = st.selectbox(
            "Snapshot Time",
            snapshot_options,
            format_func=lambda value: pd.Timestamp(value).strftime("%Y-%m-%d %H:%M"),
            key=snapshot_key,
        )
        timestamp = pd.Timestamp(raw_timestamp)
        preview = filtered.loc[filtered_times == timestamp].copy()
        preview = preview.drop(columns=["_FinderDatetime"], errors="ignore")
        st.markdown(
            "##### Middle Standard Regime Ranking — "
            f"{timestamp.strftime('%Y-%m-%d %H:%M')}"
        )
        desktop_height = min(560, 100 + 38 * len(preview))
    else:
        page_controls = st.columns([1, 1, 2])
        with page_controls[0]:
            page_size = st.selectbox(
                "Rows per page",
                [25, 50, 100, 250, 500],
                index=1,
                key=FINDER_PAGE_SIZE_STATE_KEY,
            )
        total_pages = max(1, int(np.ceil(len(filtered) / int(page_size))))
        with page_controls[1]:
            page_key = "middle_finder_page_number_20260818"
            try:
                saved_page = int(state.get(page_key, 1))
            except Exception:
                saved_page = 1
            if saved_page < 1 or saved_page > total_pages:
                state.pop(page_key, None)
            page_number = st.number_input(
                "Page",
                min_value=1,
                max_value=total_pages,
                step=1,
                key=page_key,
            )
        preview, page_meta = _finder_page(
            filtered,
            page=int(page_number),
            page_size=int(page_size),
        )
        # The historical snapshot itself may already expose ``Datetime``;
        # replace that display column before inserting the normalized Finder
        # timestamp at the front.
        preview = preview.drop(columns=["Datetime"], errors="ignore")
        preview.insert(
            0,
            "Datetime",
            _normalize_datetime_series(preview["_FinderDatetime"]).dt.strftime("%Y-%m-%d %H:%M:%S"),
        )
        preview = preview.drop(columns=["_FinderDatetime"], errors="ignore")
        with page_controls[2]:
            st.caption(
                f"Rows {page_meta['start_row']:,}–{page_meta['end_row']:,} of "
                f"{page_meta['total_rows']:,} · page {page_meta['page']:,} of "
                f"{page_meta['total_pages']:,}"
            )
        desktop_height = min(680, 100 + 38 * len(preview))

    display_columns = [column for column in preferred_columns if column in preview.columns]
    display_result = preview[display_columns] if display_columns else preview
    render_responsive_records(
        st,
        state,
        display_result,
        preferred_columns=list(display_result.columns),
        rank_column=(
            MIDDLE_AGE_RANK_COLUMN
            if MIDDLE_AGE_RANK_COLUMN in display_result.columns
            else None
        ),
        desktop_height=desktop_height,
        full_table_label="Historical Middle Standard Regime Ranking Preview",
        visualize_rank_bands=True,
        max_phone_rows=min(50, len(display_result)),
    )


def _run_finder_history_build_on_demand(state: MutableMapping[str, Any]) -> dict[str, Any]:
    """Build Finder history only after the user explicitly starts Finder."""
    try:
        from core.calculation.run_orchestrator import MARKET_RESULTS_KEY
        from core.super_quick_field3_20260722 import (
            _extract_loaded_frames,
            _fast_input_signature,
            _persist_middle_finder_history,
        )
        context = get_global_symbol_context(state)
        symbols = [str(s).strip().upper() for s in (context.loaded_symbols or context.completed_symbols) if str(s).strip()]
        frames, failures = _extract_loaded_frames(state, symbols)
        if not frames:
            return {"ok": False, "status": "NO_LOADED_FRAMES", "rows": 0}
        cutoff = min(frame["open_time"].iloc[-1] for frame in frames.values())
        aligned = {
            symbol: frame.loc[frame["open_time"] <= cutoff].copy().reset_index(drop=True)
            for symbol, frame in frames.items()
            if isinstance(frame, pd.DataFrame) and not frame.empty
        }
        signature = _fast_input_signature(
            aligned,
            selected=symbols,
            timeframe=str(context.timeframe or state.get("timeframe") or "H1").upper(),
            failures=failures,
        )
        return _persist_middle_finder_history(
            state,
            aligned,
            timeframe=str(context.timeframe or state.get("timeframe") or "H1").upper(),
            signature=signature,
        )
    except Exception as exc:
        return {"ok": False, "status": "FINDER_BUILD_FAILED", "rows": 0, "error": f"{type(exc).__name__}: {exc}"}


# ============================================================================
# Current Middle download
# ============================================================================

def render_middle_regime_download_button(
    st: Any,
    middle_age: pd.DataFrame,
) -> None:

    if (
        not isinstance(
            middle_age,
            pd.DataFrame,
        )
        or middle_age.empty
    ):
        return

    middle_age = _canonical_middle_export_frame(middle_age)

    # ----------------------------------------------------------------------
    # CSV — always render this independently of optional Excel support.
    # ----------------------------------------------------------------------

    try:
        csv = middle_age.to_csv(
            index=False,
            encoding="utf-8",
        )
    except Exception as exc:
        st.error(f"Middle Standard CSV could not be prepared: {type(exc).__name__}: {exc}")
        csv = None

    if csv is not None:

        st.download_button(
            label=(
                "📥 Download Middle Standard Regime Ranking CSV"
            ),
            data=csv,
            file_name=(
                "middle_standard_regime_ranking.csv"
            ),
            mime="text/csv",
            key=(
                "download_middle_standard_regime_csv"
            ),
            use_container_width=True,
        )

    # ----------------------------------------------------------------------
    # Colored Excel
    # ----------------------------------------------------------------------

    try:
        xlsx = _dataframe_to_excel(
            middle_age,
            "Middle Regime Ranking",
        )
    except Exception as exc:
        state_error = f"{type(exc).__name__}: {exc}"
        xlsx = None
        st.warning(f"Colored Excel is unavailable for this render: {state_error}")

    if xlsx is not None:

        st.download_button(
            label=(
                "🎨 Download Middle Standard Regime Colored Excel"
            ),
            data=xlsx,
            file_name=(
                "middle_standard_regime_ranking.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            key=(
                "download_middle_standard_regime_xlsx"
            ),
            use_container_width=True,
        )


# ============================================================================
# Main render
# ============================================================================

def render_multisymbol_regime_summary(
    st: Any,
    state: MutableMapping[str, Any],
) -> None:

    _load_saved_if_needed(
        state
    )

    context = (
        get_global_symbol_context(
            state
        )
    )

    loaded = list(
        context.completed_symbols
        or context.loaded_symbols
    )

    raw_ranking = _sanitize_index_column_collision(
        state.get("field3_multisymbol_regime_20260708")
    )

    raw_evidence = _sanitize_index_column_collision(
        state.get("field3_regime_evidence_v2")
    )

    if (
        not isinstance(
            raw_ranking,
            pd.DataFrame,
        )
        or raw_ranking.empty
    ):

        st.info(
            "No completed multi-symbol Field 3 generation is available. "
            "Select and load symbols in Settings, then run Super Quick, "
            "Quick, or Full Calculation."
        )

        return

    full_evidence = enrich_evidence(
        raw_evidence
        if isinstance(
            raw_evidence,
            pd.DataFrame,
        )
        else pd.DataFrame()
    )

    full_ranking = _sanitize_index_column_collision(
        enrich_ranking(
            raw_ranking,
            full_evidence,
        )
    )
    full_ranking = _fill_middle_starting_prices_from_loaded_candles(
        full_ranking,
        state,
    )

    # Publish canonical enriched data.
    state[
        "field3_multisymbol_regime_20260708"
    ] = full_ranking

    state[
        "field3_regime_evidence_v2"
    ] = full_evidence

    ranking = _scope_to_loaded(
        full_ranking,
        loaded,
    )

    evidence = _scope_to_loaded(
        full_evidence,
        loaded,
    )

    # =========================================================================
    # Header
    # =========================================================================

    st.markdown(
        "## All Loaded Symbols — Three Standards and Final Ranking"
    )

    st.caption(
        "Aggregate Field 3 universe. The active Global Symbol "
        "does not alter this section."
    )

    phone_mode = is_phone_mode(
        state
    )

    metrics = st.columns(
        2 if phone_mode else 5
    )

    displayed_timeframe = (
        context.timeframe
        or (
            str(
                ranking[
                    "Timeframe"
                ].iloc[0]
            )
            if (
                not ranking.empty
                and "Timeframe"
                in ranking.columns
            )
            else "—"
        )
    )

    if phone_mode:

        metrics[0].metric(
            "Loaded",
            (
                len(loaded)
                if loaded
                else len(ranking)
            ),
        )

        metrics[1].metric(
            "Timeframe",
            displayed_timeframe,
        )

    else:

        metrics[0].metric(
            "Loaded / Completed",
            (
                len(loaded)
                if loaded
                else len(ranking)
            ),
        )

        metrics[1].metric(
            "Ranking Rows",
            len(ranking),
        )

        metrics[2].metric(
            "Standard Rows",
            len(evidence),
        )

        metrics[3].metric(
            "Timeframe",
            displayed_timeframe,
        )

        metrics[4].metric(
            "Generation",
            context.generation or "—",
        )

    if loaded:

        st.caption(
            "Aggregate universe: "
            + " → ".join(
                loaded
            )
        )

    # =========================================================================
    # Canonical standard tables
    #
    # Build Middle and Higher from the exact same enriched ranking before any
    # Open/Close UI is entered.  This prevents an expander interaction from
    # replacing one table with a separately scoped or partially rebuilt frame.
    # =========================================================================
    middle_age = _middle_age_ranking(
        ranking
    )
    middle_age = _add_weighted_risk_score(middle_age, state)

    higher_age = _age_only_ranking(
        ranking
    )

    canonical_tables = {
        "ranking": full_ranking.copy(),
        "evidence": full_evidence.copy(),
        "middle": middle_age.copy(),
        "higher": higher_age.copy(),
    }
    state[FIELD3_CANONICAL_TABLES_KEY] = canonical_tables
    _publish_higher_ranking_to_field_sections(
        state,
        higher_age,
    )

    # Do not load/merge/write the potentially very large Finder history while
    # rendering the instant Field 3 result. Finder owns that work and starts it
    # only from its explicit Run Finder button below.

    # =========================================================================
    # Current Middle table and Finder are sibling sections.  Both consume the
    # canonical ``middle`` frame/history above, so either section can be
    # opened without disabling the other.
    # =========================================================================

    st.markdown(
        "### Middle Standard Regime Ranking"
    )

    if middle_age.empty:
        st.info("Middle Standard Regime ranking is not available.")
    else:
        render_middle_regime_download_button(st, middle_age)
        st.caption(
            "Risk_Score = MS × CC ÷ SA. Scores ≥ 1.0 are potential blocks. "
            "The rolling hourly quota blocks the first four potential blocks "
            "and allows later ones after the quota is reached. "
            "Symbols are prioritized from youngest SA to oldest. SL is No SL "
            "when SA is below 15 and Allowed otherwise."
        )
        render_responsive_records(
            st,
            state,
            middle_age,
            preferred_columns=[
                MIDDLE_AGE_RANK_COLUMN,
                SL_DECISION_COLUMN,
                "Symbol",
                STRATEGY_DECISION_COLUMN,
                "Lower Regime Bias",
                "Middle Regime Bias",
                "Higher Standard Regime Bias",
                CANDLE_AGE_COLUMN,
                STARTING_PRICE_COLUMN,
                RISK_SCORE_COLUMN,
                IS_POTENTIAL_BLOCK_COLUMN,
                HOURLY_BLOCK_COUNT_COLUMN,
                FINAL_ACTION_COLUMN,
                "Hour",
                WEIGHTED_RISK_SCORE_COLUMN,
                WEIGHTED_RISK_DECISION_COLUMN,
                RAS_DECISION_COLUMN,
                CANDLE_STATISTICS_COLUMN,
                AVG_REGIME_CANDLE_COLUMN,
                COMPLETED_SAMPLE_DISPLAY_COLUMN,
                "Middle Standard Regime State",
                "Lower Standard Regime State",
                "Higher Standard Regime State",
            ],
            rank_column=MIDDLE_AGE_RANK_COLUMN,
            desktop_height=min(620, 100 + 38 * len(middle_age)),
            full_table_label="Full Middle Standard Regime Ranking",
            visualize_rank_bands=True,
            match_columns=("Middle Regime Bias", "Higher Standard Regime Bias"),
            highlight_match_row=True,
            cell_extreme_highlights=MIDDLE_NUMERIC_HIGHLIGHT_COLUMNS,
            highlight_min_cells=True,
            highlight_max_cells=True,
            max_phone_rows=30,
            force_table_on_phone=True,
            column_config={
                STRATEGY_DECISION_COLUMN: st.column_config.TextColumn(
                    STRATEGY_DECISION_COLUMN,
                    width="large",
                    help="Direction first, then strategy level: BUY/SELL — S1/S2/S3/S4/None.",
                ),
            },
        )

    with st.expander(
        "🔎 Middle Regime Historical Finder",
        expanded=False,
    ):
        finder_running = bool(state.get("field3_middle_finder_run_requested_20260824"))
        if not finder_running:
            st.info(
                "Finder is idle. Super Quick does not build Finder history, so the "
                "Middle Standard ranking above can appear immediately."
            )
            if st.button(
                "🔎 Run Finder",
                key="field3_middle_finder_run_button_20260824",
                use_container_width=True,
                help="Build/search historical Middle Regime rows from the already-loaded candles.",
            ):
                state["field3_middle_finder_run_requested_20260824"] = True
                state["field3_middle_finder_run_started_20260824"] = True
                st.rerun()
        else:
            close_finder_col, finder_status_col = st.columns([1, 3])
            if close_finder_col.button(
                "Close Finder",
                key="field3_middle_finder_close_button_20260824",
                use_container_width=True,
            ):
                state["field3_middle_finder_run_requested_20260824"] = False
                st.rerun()
            if not state.get("field3_middle_finder_history_ready_20260824"):
                with st.spinner("Building Finder history from loaded candles…"):
                    finder_build = _run_finder_history_build_on_demand(state)
                state["field3_middle_finder_history_ready_20260824"] = bool(finder_build.get("ok"))
                state["field3_middle_finder_history_build_20260824"] = finder_build
                if not finder_build.get("ok"):
                    finder_status_col.warning(
                        "Finder could not build history yet: "
                        f"{finder_build.get('error') or finder_build.get('status') or 'NO_DATA'}"
                    )
            try:
                render_middle_regime_finder(st, state)
            except Exception as exc:
                state["field3_middle_finder_render_error_20260818"] = (
                    f"{type(exc).__name__}: {exc}"
                )
                st.error(
                    "Middle Regime Historical Finder could not render safely. "
                    "The current Middle Standard ranking remains available above: "
                    f"{type(exc).__name__}: {exc}"
                )

    # =========================================================================
    # Standard tables
    # =========================================================================

    fast_two_table_mode = bool(
        state.get(
            "field3_fast_two_table_mode_20260722"
        )
    ) or (
        str(
            state.get(
                "field3_last_run_scope_20260722"
            )
            or ""
        ).upper()
        == "LUNCH_CORE"
    )

    # Higher Standard Summary is an internal evidence source only.  The user
    # requested that its standalone summary and the duplicate Field Open/Close
    # Higher ranking panels not be rendered in the Field 3 page.
    standards_to_render = tuple(STANDARD_LABELS.items())

    for standard, label in (
        standards_to_render
    ):

        st.markdown(
            f"### {label}"
        )

        table = _standard_table(
            evidence,
            ranking,
            standard,
        )

        if table.empty:

            st.info(
                f"No saved {standard.title()} evidence is available."
            )

        else:

            render_responsive_records(
                st,
                state,
                table,
                preferred_columns=[
                    RECENT_CHANGE_RANK_COLUMN,
                    "Symbol",
                    "Regime State",
                    "Bias",
                    CANDLE_AGE_COLUMN,
                    "Calibrated Reliability",
                    "Posterior Probability",
                    "Transition Risk",
                    "Data Quality Grade",
                ],
                rank_column=(
                    RECENT_CHANGE_RANK_COLUMN
                ),
                desktop_height=min(
                    520,
                    92 + 38 * len(
                        table
                    ),
                ),
                full_table_label=(
                    f"Full {label} table"
                ),
            )

    if fast_two_table_mode:

        return

    # =========================================================================
    # Final ranking
    # =========================================================================

    st.markdown(
        "### Final Cross-Symbol Ranking"
    )

    final_columns = [
        "Rank",
        REGIME_START_AGE_RANK_COLUMN,
        RECENT_CHANGE_RANK_COLUMN,
        "Symbol",
        "Lower Regime",
        "Lower Bias",
        "Lower Probability",
        "Lower Reliability",
        "Middle Regime",
        "Middle Bias",
        "Middle Probability",
        "Middle Reliability",
        "Higher Regime",
        "Higher Bias",
        "Higher Probability",
        "Higher Reliability",
        CANDLE_AGE_COLUMN,
        REGIME_START_STANDARD_COLUMN,
        "Three-Regime Agreement",
        "Composite Bias",
        "Composite Score",
        "Decision Strength",
        "Calibrated Reliability",
        "Entry Permission",
        "Block Reason",
        "Completed Candle",
    ]

    final = ranking[
        [
            c
            for c in final_columns
            if c in ranking.columns
        ]
    ].copy()

    if "Rank" in final.columns:

        final = final.sort_values(
            "Rank",
            kind="mergesort",
        )

    render_responsive_records(
        st,
        state,
        final,
        preferred_columns=[
            "Rank",
            "Symbol",
            "Higher Regime",
            "Higher Bias",
            CANDLE_AGE_COLUMN,
            "Composite Bias",
            "Composite Score",
            "Decision Strength",
            "Calibrated Reliability",
            "Entry Permission",
            "Block Reason",
        ],
        rank_column="Rank",
        desktop_height=min(
            720,
            100 + 38 * len(
                final
            ),
        ),
        full_table_label=(
            "Full Cross-Symbol Ranking table"
        ),
        visualize_rank_bands=True,
        match_columns=(
            "Higher Regime",
            "Middle Regime",
        ),
        max_phone_rows=30,
    )


# ============================================================================
# Public exports
# ============================================================================

__all__ = [
    "render_multisymbol_regime_summary",
    "render_middle_regime_finder",
    "render_middle_regime_download_button",

    "render_field_close_higher_ranking",
    "render_field_open_higher_ranking",

    "_age_only_ranking",
    "_middle_age_ranking",

    "get_ras_status",
    "get_strategy_decision",

    "_add_ras_decision_column",
    "_add_strategy_decision_column",
    "_add_weighted_risk_score",
    "_add_hourly_capped_risk_filter",

    "AGE_ONLY_RANK_COLUMN",
    "HIGHER_AGE_RANK_COLUMN",
    "HIGHER_AGE_DISPLAY_COLUMN",
    "FIELD3_CANONICAL_TABLES_KEY",
    "MIDDLE_AGE_RANK_COLUMN",
    "MIDDLE_AGE_DISPLAY_COLUMN",
    "MIDDLE_AVG_SOURCE_COLUMN",
    "AVG_REGIME_CANDLE_COLUMN",
    "MIDDLE_COMPLETED_SAMPLE_COLUMN",
    "COMPLETED_SAMPLE_DISPLAY_COLUMN",
    "CANDLE_STATISTICS_COLUMN",
    "RISK_SCORE_COLUMN",
    "IS_POTENTIAL_BLOCK_COLUMN",
    "HOURLY_BLOCK_COUNT_COLUMN",
    "FINAL_ACTION_COLUMN",
    "RISK_SCORE_THRESHOLD",
    "HOURLY_BLOCK_QUOTA",
    "WEIGHTED_RISK_SCORE_COLUMN",
    "WEIGHTED_RISK_DECISION_COLUMN",
    "WEIGHTED_RISK_THRESHOLD",
    "SL_DECISION_COLUMN",

    "RAS_DECISION_COLUMN",
    "STRATEGY_DECISION_COLUMN",
]
