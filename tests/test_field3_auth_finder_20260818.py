"""Focused regression checks for the Field 3 reconnect and Finder repairs.

Run directly with:
    python tests/test_field3_auth_finder_20260818.py

The small import stubs keep this test runnable in packaging environments that
do not install the full Streamlit UI dependency stack.
"""
from __future__ import annotations

import importlib.util
import os
from pathlib import Path
import sys
import tempfile
import types

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


def _load_finder_module():
    regime_stub = types.ModuleType("core.regime_age_columns_20260722")
    regime_stub.CANDLE_AGE_COLUMN = "Candle After Regime Start"
    regime_stub.RECENT_CHANGE_RANK_COLUMN = "Recent Regime Change Rank"
    regime_stub.REGIME_START_AGE_RANK_COLUMN = "Regime Start Age Rank"
    regime_stub.REGIME_START_STANDARD_COLUMN = "Regime Start Standard"
    regime_stub.STANDARD_AGE_COLUMNS = {
        "LOWER": "Lower Candle After Regime Start",
        "MIDDLE": "Middle Candle After Regime Start",
        "HIGHER": "Higher Candle After Regime Start",
    }
    regime_stub.enrich_evidence = lambda frame, *_args, **_kwargs: frame
    regime_stub.enrich_ranking = lambda frame, *_args, **_kwargs: frame
    sys.modules[regime_stub.__name__] = regime_stub
    return _load_module(
        "field3_finder_under_test",
        ROOT / "ui" / "field3_multisymbol_regime_summary_20260722.py",
    )


def test_reconnect_opens_field3_without_login() -> None:
    class FakeStreamlit(types.ModuleType):
        def __init__(self):
            super().__init__("streamlit")
            self.session_state = {}
            self.query_params = {}

    fake_st = FakeStreamlit()
    sys.modules["streamlit"] = fake_st
    auth = _load_module(
        "auth_under_test",
        ROOT / "core" / "light_auth_20260612.py",
    )

    with tempfile.TemporaryDirectory() as temporary:
        old_db = os.environ.get("NEW7_AUTH_DB")
        os.environ["NEW7_AUTH_DB"] = str(Path(temporary) / "auth.sqlite3")
        try:
            auth._login_success("researcher@example.com", guest=False)
            token = fake_st.query_params.get(auth.AUTH_TOKEN_QUERY_KEY)
            assert token and "researcher@example.com" not in token

            auth.request_persistent_auth_route("Field 3")
            fake_st.session_state = {}
            assert auth.render_auth_gate() is True
            assert fake_st.session_state["new7_auth_logged_in"] is True
            assert fake_st.session_state["active_page"] == "Field 3"
            assert fake_st.session_state["field_3_expanded"] is True
            assert auth.AUTH_TARGET_QUERY_KEY not in fake_st.query_params

            auth.clear_persistent_auth_session()
            fake_st.session_state = {}
            fake_st.query_params[auth.AUTH_TOKEN_QUERY_KEY] = token
            assert auth._restore_reconnect_session(auth._db_path(), token) is None
        finally:
            if old_db is None:
                os.environ.pop("NEW7_AUTH_DB", None)
            else:
                os.environ["NEW7_AUTH_DB"] = old_db


def test_finder_range_filters_pages_and_exports() -> None:
    finder = _load_finder_module()
    times = pd.to_datetime(
        [
            "2026-08-01 10:00:00",
            "2026-08-01 10:00:00",
            "2026-08-01 11:00:00",
            "2026-08-01 11:00:00",
            "2026-08-01 12:00:00",
            "2026-08-01 12:00:00",
        ]
    )
    history = pd.DataFrame(
        {
            "_FinderDatetime": times,
            "Symbol": ["EURUSD", "GBPUSD"] * 3,
            finder.STRATEGY_DECISION_COLUMN: ["S1", "S2", "S1", "S2", "S3", "S4"],
            finder.MIDDLE_AGE_RANK_COLUMN: [1, 2, 1, 2, 1, 2],
            "Score": np.arange(6),
        }
    )

    result = finder._finder_range(
        history,
        pd.Timestamp("2026-08-01 10:00:00"),
        pd.Timestamp("2026-08-01 12:00:00"),
    )
    assert len(result) == 6, "Both exact endpoints must be included."

    filtered = finder._filter_finder_result(result, symbols=["eurusd"], strategies=["S1"])
    assert len(filtered) == 2
    assert set(filtered["Symbol"]) == {"EURUSD"}

    page, meta = finder._finder_page(result, page=2, page_size=10)
    assert len(page) == 6
    assert meta == {
        "page": 1,
        "page_size": 10,
        "total_pages": 1,
        "total_rows": 6,
        "start_row": 1,
        "end_row": 6,
    }

    csv_text = finder._finder_csv(
        result,
        pd.Timestamp("2026-08-01 10:00:00"),
        pd.Timestamp("2026-08-01 12:00:00"),
    )
    assert "2026-08-01 10:00:00" in csv_text
    assert "2026-08-01 12:00:00" in csv_text
    assert csv_text.count("\n") == 7

    workbook_bytes = finder._finder_to_excel(
        result,
        pd.Timestamp("2026-08-01 10:00:00"),
        pd.Timestamp("2026-08-01 12:00:00"),
    )
    assert workbook_bytes
    from io import BytesIO

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)
    assert workbook.sheetnames == ["Finder Results", "Range"]
    assert workbook["Finder Results"].max_row == 7
    assert workbook["Range"]["B5"].value == 6



def test_middle_history_handles_datetime_index_column_collision() -> None:
    finder = _load_finder_module()
    frame = pd.DataFrame(
        {
            "Datetime": pd.to_datetime(["2026-08-01 10:00:00"]),
            "Symbol": ["EURUSD"],
            finder.MIDDLE_AGE_RANK_COLUMN: [1],
        }
    )
    frame.index = pd.DatetimeIndex(frame["Datetime"], name="Datetime")
    safe = finder._sanitize_index_column_collision(frame)
    assert safe.index.name is None
    assert "Datetime" in safe.columns
    assert safe.reset_index(drop=True).shape == (1, 3)


def test_router_targets_field3() -> None:
    source = (ROOT / "tabs" / "antd_page_router_20260615.py").read_text(encoding="utf-8")
    function_source = source[source.index("def _open_lunch_ai_after_settings_run"):]
    function_source = function_source[: function_source.index("\ndef _render_arert_research_settings")]
    assert 'request_persistent_auth_route("Field 3")' in function_source
    assert 'number == 3' in function_source
    assert 'number == 10' not in function_source


def test_super_quick_builds_finder_history_for_every_symbol() -> None:
    orchestrator_stub = types.ModuleType("core.calculation.run_orchestrator")
    orchestrator_stub.MARKET_RESULTS_KEY = "market_results"
    sys.modules[orchestrator_stub.__name__] = orchestrator_stub

    engine_stub = types.ModuleType("core.field3_three_regime_engine")
    engine_stub.standard_windows = lambda _timeframe: {
        "LOWER": 24,
        "MIDDLE": 120,
        "HIGHER": 600,
    }

    def standardize(frame: pd.DataFrame) -> pd.DataFrame:
        output = frame.rename(columns={"Datetime": "open_time"}).copy()
        output["open_time"] = pd.to_datetime(output["open_time"], errors="coerce", utc=True)
        return output.dropna(subset=["open_time"]).sort_values("open_time").reset_index(drop=True)

    engine_stub.standardize_candles = standardize
    sys.modules[engine_stub.__name__] = engine_stub
    super_quick = _load_module(
        "super_quick_under_test",
        ROOT / "core" / "super_quick_field3_20260722.py",
    )

    row_count = 160
    datetimes = pd.date_range("2026-01-01", periods=row_count, freq="h")
    close = 1.1 + np.linspace(0, 0.04, row_count) + np.sin(np.arange(row_count) / 8) * 0.002
    first = pd.DataFrame(
        {
            "Datetime": datetimes,
            "open": close - 0.0002,
            "high": close + 0.0005,
            "low": close - 0.0005,
            "close": close,
            "volume": np.arange(row_count) + 100,
        }
    )
    second = first.assign(
        open=(close - 0.0002) * 1.12,
        high=(close + 0.0005) * 1.12,
        low=(close - 0.0005) * 1.12,
        close=close * 1.12,
    )
    history = super_quick._build_middle_finder_history(
        {"EURUSD": first, "GBPUSD": second},
        timeframe="H1",
    )
    assert not history.empty
    assert set(history["Symbol"]) == {"EURUSD", "GBPUSD"}
    assert history.duplicated(["Datetime", "Symbol"]).sum() == 0
    assert history["Datetime"].is_monotonic_increasing


def test_middle_finder_and_higher_tables_share_one_publication() -> None:
    """The Finder, Middle table, and both Higher Open/Close views stay linked."""
    finder = _load_finder_module()
    ranking = pd.DataFrame(
        {
            "Symbol": ["EURUSD", "GBPUSD"],
            "Lower Candle After Regime Start": [3, 6],
            "Middle Candle After Regime Start": [5, 8],
            "Higher Candle After Regime Start": [7, 11],
            "Lower Regime": ["BULL", "BEAR"],
            "Lower Bias": ["BUY", "SELL"],
            "Middle Regime": ["BULL", "BEAR"],
            "Middle Bias": ["BUY", "SELL"],
            "Higher Regime": ["BULL", "BEAR"],
            "Higher Bias": ["BUY", "SELL"],
            "Completed Candle": pd.to_datetime(["2026-08-01", "2026-08-01"]),
            "Timeframe": ["H1", "H1"],
        }
    )
    middle = finder._middle_age_ranking(ranking)
    higher = finder._age_only_ranking(ranking)
    assert not middle.empty
    assert not higher.empty
    assert set(higher["Symbol"]) == set(middle["Symbol"])
    assert "Higher Regime Age Rank" in higher.columns
    assert "RAS Decision" in higher.columns
    assert "Strategy Decision" in higher.columns

    state = {}
    finder._publish_higher_ranking_to_field_sections(state, higher)
    assert state[finder.FIELD_CLOSE_HIGHER_RANKING_KEY].equals(
        state[finder.FIELD_OPEN_HIGHER_RANKING_KEY]
    )
    assert state[finder.FIELD_CLOSE_HIGHER_MOBILE_KEY].equals(
        state[finder.FIELD_OPEN_HIGHER_MOBILE_KEY]
    )

    history = middle.copy()
    history["Datetime"] = pd.to_datetime(
        ["2026-08-01 10:00", "2026-08-01 10:00"]
    )
    prepared = finder._prepare_middle_history(history)
    result = finder._finder_range(
        prepared,
        pd.Timestamp("2026-08-01 10:00"),
        pd.Timestamp("2026-08-01 10:00"),
    )
    assert set(result["Symbol"]) == set(middle["Symbol"])


if __name__ == "__main__":
    test_reconnect_opens_field3_without_login()
    test_finder_range_filters_pages_and_exports()
    test_middle_history_handles_datetime_index_column_collision()
    test_router_targets_field3()
    test_super_quick_builds_finder_history_for_every_symbol()
    test_middle_finder_and_higher_tables_share_one_publication()
    print(
        "PASS: Field 3 reconnect route, Finder range/filter/page/export, "
        "router target, Super Quick history backfill, and joint Middle/Higher publication"
    )
